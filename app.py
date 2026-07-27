from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
from datetime import timedelta, datetime
import sqlite3
import os
import psycopg2
import psycopg2.extras
import json
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import tempfile
from reportlab.lib.pagesizes import A4, landscape
import requests
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from supabase import create_client
import uuid
from dotenv import load_dotenv

load_dotenv()


app = Flask(__name__)
app.secret_key = "sai-fuel-mart-secret-key"
app.permanent_session_lifetime = timedelta(days=30)


@app.route("/service-worker.js")
def service_worker():
    """
    Served from the root (not /static/) so its control scope covers the
    whole app — a service worker registered from /static/service-worker.js
    would only be allowed to control pages under /static/, not the actual
    app pages like /dashboard or /tank-level.
    """
    return app.send_static_file("service-worker.js")

USERS = {

    "admin": {
        "password": "1234",
        "role": "admin"
    },

    "manager": {
        "password": "1234",
        "role": "manager"
    }

}

DB_PATH = os.path.join(os.getcwd(), "sai_fuel_mart.db")

def is_admin():
    return session.get("role") == "admin"

def is_manager():
    return session.get("role") == "manager"

def get_conn():
    conn = sqlite3.connect(
        DB_PATH,
        timeout=30,
        check_same_thread=False
    )
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn

def style_excel_sheet(ws):
    header_fill = PatternFill(start_color="07120C", end_color="07120C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for column_cells in ws.columns:
        max_length = 0
        col_no = column_cells[0].column

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col_no)].width = max_length + 4

def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            id INTEGER PRIMARY KEY,
            ms_rate REAL,
            hsd_rate REAL,
            cng_rate REAL
        )
    """)

    cur.execute("""
        INSERT OR IGNORE INTO settings (id, ms_rate, hsd_rate, cng_rate)
        VALUES (1, 102.46, 93.72, 78)
    """)

    try:
        cur.execute("ALTER TABLE settings ADD COLUMN cng_rate REAL DEFAULT 0")
    except sqlite3.OperationalError:
        pass

    cur.execute("""
        CREATE TABLE IF NOT EXISTS daily_closing (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            ms_litres REAL,
            hsd_litres REAL,
            cng_sale REAL,
            total_fuel_sale REAL,
            lube_sale REAL,
            digital_collection REAL,
            credit_given REAL,
            transport_received REAL,
            net_credit_due REAL,
            total_expense REAL,
            cash_in_hand REAL
        )
    """)

    

    for col, typ in [
        ("phonepe", "REAL DEFAULT 0"),
        ("card_swipe", "REAL DEFAULT 0"),
        ("hp_pay", "REAL DEFAULT 0"),
        ("hpcl_otp", "REAL DEFAULT 0"),
        ("upi_other", "REAL DEFAULT 0")
    ]:
        try:
            cur.execute(f"ALTER TABLE daily_closing ADD COLUMN {col} {typ}")
        except sqlite3.OperationalError:
            pass

    try:
        cur.execute("ALTER TABLE daily_closing ADD COLUMN cng_sale REAL DEFAULT 0")
    except sqlite3.OperationalError:
        pass

     

    cur.execute("""
        CREATE TABLE IF NOT EXISTS tank_level (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            fuel_type TEXT,
            opening_stock REAL,
            received_stock REAL,
            own_tanker_stock REAL,
            sale_stock REAL,
            gain_qty REAL,
            shortage_qty REAL,
            current_stock REAL,
            gain_amount REAL,
            shortage_amount REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
CREATE TABLE IF NOT EXISTS salary_payments(

    id INTEGER PRIMARY KEY AUTOINCREMENT,

    payment_date TEXT,

    emp_id TEXT,

    employee_name TEXT,

    payment_mode TEXT,

    bank_account TEXT,

    amount REAL,

    month_name TEXT,

    remarks TEXT

)
""")

    cur.execute("""
        CREATE TABLE IF NOT EXISTS staff_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_name TEXT,
            role TEXT,
            status TEXT
        )
    """)

    for col, typ in [
    ("emp_id", "TEXT"),
    ("department", "TEXT"),
    ("joined_date", "TEXT"),
    ("bank_account", "TEXT"),
    ("shift", "TEXT")
    ]:
        try:
            cur.execute(f"ALTER TABLE staff_master ADD COLUMN {col} {typ}")
        except sqlite3.OperationalError:
            pass

        
      
    cur.execute("""
    CREATE TABLE IF NOT EXISTS credit_transporters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        party_name TEXT,
        opening_balance REAL DEFAULT 0,
        credit_given REAL DEFAULT 0,
        payment_received REAL DEFAULT 0,
        balance_due REAL DEFAULT 0,
        status TEXT
    )
""")

    for col, typ in [

    ("mobile", "TEXT"),
    ("vehicle_no", "TEXT"),

    ("fuel_credit", "REAL DEFAULT 0"),
    ("lube_credit", "REAL DEFAULT 0"),

    ("amount_received", "REAL DEFAULT 0")

]:
        try:
            cur.execute(
            f"ALTER TABLE credit_transporters ADD COLUMN {col} {typ}"
        )
        except sqlite3.OperationalError:
         pass

     
    
    try:
       cur.execute(
        "ALTER TABLE transport_entries ADD COLUMN transporter_id INTEGER"
    )
    except sqlite3.OperationalError:
     pass
    cur.execute("""

    CREATE TABLE IF NOT EXISTS transporter_ledger (

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        date TEXT,

        transporter_id INTEGER,
        transporter_name TEXT,

        entry_type TEXT,

        fuel_credit REAL DEFAULT 0,
        lube_credit REAL DEFAULT 0,

        received_amount REAL DEFAULT 0,

        balance_after REAL DEFAULT 0,

        remarks TEXT,

        created_at TEXT DEFAULT CURRENT_TIMESTAMP

    )

""") 
    cur.execute("""
    CREATE TABLE IF NOT EXISTS lube_transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        product_id INTEGER,
        product_name TEXT,
        transaction_type TEXT,
        qty REAL,
        rate REAL,
        amount REAL,
        remarks TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
""")

    cur.execute("""
        CREATE TABLE IF NOT EXISTS transport_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            transporter_id INTEGER,
            payment_amount REAL
        )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS credit_lube_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        transporter_id INTEGER,
        transporter_name TEXT,
        product_id INTEGER,
        product_name TEXT,
        qty REAL,
        rate REAL,
        amount REAL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
""")

    cur.execute("""
        CREATE TABLE IF NOT EXISTS transport_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_date TEXT,
            sl_no INTEGER,
            transporter_name TEXT,
            challan_no TEXT,
            vehicle_no TEXT,
            slip_no TEXT,
            hsd_qty REAL,
            rate REAL,
            hsd_amount REAL,
            cash_taken REAL,
            total_amount REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    try:
        cur.execute("ALTER TABLE transport_entries ADD COLUMN transporter_name TEXT")
    except sqlite3.OperationalError:
        pass

    cur.execute("""
        CREATE TABLE IF NOT EXISTS nozzle_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nozzle_name TEXT,
            machine_no TEXT,
            fuel_type TEXT,
            status TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS nozzle_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_date TEXT,
            nozzle_id INTEGER,
            opening_reading REAL,
            closing_reading REAL,
            testing_qty REAL,
            total_sale REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    default_nozzles = [

    ("MS1", "351", "MS", "Active"),
    ("MS2", "356", "MS", "Active"),
    ("MS3", "725", "MS", "Active"),

    ("HSD1", "351", "HSD", "Active"),
    ("HSD2", "356", "HSD", "Active"),
    ("HSD3", "725", "HSD", "Active"),

    ("CNG1", "CNG-1", "CNG", "Active"),
    ("CNG2", "CNG-2", "CNG", "Active")

]

    for nozzle in default_nozzles:
        cur.execute("SELECT id FROM nozzle_master WHERE nozzle_name = ?", (nozzle[0],))
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO nozzle_master (
                    nozzle_name, machine_no, fuel_type, status
                )
                VALUES (?, ?, ?, ?)
            """, nozzle)

    conn.commit()
    conn.close()


@app.route("/")
def home():
    if session.get("logged_in"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():

    error = ""

    if request.method == "POST":

        username = request.form.get("username")
        password = request.form.get("password")

        # CHECK USER

        if username in USERS and USERS[username]["password"] == password:

            session["logged_in"] = True

            session["username"] = username

            session["role"] = USERS[username]["role"]

            # REMEMBER LOGIN

            if request.form.get("remember"):
                session.permanent = True

            return redirect(url_for("dashboard"))

        else:

            error = "Invalid username or password"

    return render_template(
        "login.html",
        error=error
    )


@app.route("/dashboard")
def dashboard():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM daily_closing ORDER BY id DESC LIMIT 1")
    latest = cur.fetchone()

    cur.execute("""

    SELECT

        ROUND(
            SUM(total_fuel_sale),
            2
        ) AS monthly_sale,

        ROUND(
            SUM(cng_sale * (
                SELECT cng_rate
                FROM settings
                WHERE id=1
            )),
            2
        ) AS monthly_cng,

        ROUND(
            SUM(total_expense),
            2
        ) AS monthly_expense,

        ROUND(
            SUM(net_credit_due),
            2
        ) AS monthly_credit

    FROM daily_closing

    WHERE TO_CHAR(date, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')

""")
    monthly = cur.fetchone()

    cur.execute("""
    SELECT
        dc.date,
        COALESCE(dc.total_fuel_sale, 0) AS total_fuel_sale,
        COALESCE(dc.cng_sale, 0) * COALESCE(s.cng_rate, 0) AS cng_revenue
    FROM daily_closing dc
    CROSS JOIN settings s
    WHERE s.id = 1
    ORDER BY dc.date ASC
    LIMIT 15
""") 
    chart_rows = cur.fetchall()

    chart_labels = []
    chart_sales = []
    cng_data = []

    for r in chart_rows:
        chart_labels.append(r["date"])
        chart_sales.append(round(float(r["total_fuel_sale"] or 0)))

        if "cng_revenue" in r.keys():
            cng_data.append(round(float(r["cng_revenue"] or 0)))
        else:
            cng_data.append(round(float(r["cng_sale"] or 0)))

    cur.execute("""
        SELECT party_name, credit_given, payment_received, balance_due
        FROM credit_transporters
        ORDER BY balance_due DESC
    """)
    transporter_summary = cur.fetchall()

    cur.execute("""
        SELECT product_name, closing_stock
        FROM lube_stock
        WHERE closing_stock <= 5
        ORDER BY closing_stock ASC
    """)
    low_lube = cur.fetchall()

    cur.execute("SELECT * FROM tank_level WHERE fuel_type='MS' ORDER BY id DESC LIMIT 1")
    ms_tank = cur.fetchone()

    cur.execute("SELECT * FROM tank_level WHERE fuel_type='HSD' ORDER BY id DESC LIMIT 1")
    hsd_tank = cur.fetchone()

    today = datetime.now()
    show_backup_popup = today.day <= 3

    conn.close()

    return render_template(
        "dashboard.html",
        chart_labels=chart_labels,
        chart_sales=chart_sales,
        cng_data=cng_data,
        latest=latest,
        monthly=monthly,
        chart_rows=chart_rows,
        transporter_summary=transporter_summary,
        low_lube=low_lube,
        ms_tank=ms_tank,
        hsd_tank=hsd_tank,
        show_backup_popup=show_backup_popup
    )

@app.route("/receive-transport-payment/<int:id>", methods=["POST"])
def receive_transport_payment(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    amount = round(float(request.form.get("amount", 0) or 0), 2)

    if amount <= 0:
        return redirect(url_for("credit_transport"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE credit_transporters
            SET
                payment_received = COALESCE(payment_received,0) + %s,
                balance_due = COALESCE(balance_due,0) - %s
            WHERE id=%s
            RETURNING *
        """, (
            amount,
            amount,
            id
        ))

        party = cur.fetchone()

        if party:
            cur.execute("""
                INSERT INTO transporter_ledger(
                    date, transporter_id, transporter_name, entry_type,
                    fuel_credit, lube_credit, received_amount, balance_after, remarks
                )
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                datetime.now().strftime("%Y-%m-%d"),
                id,
                party["party_name"],
                "Payment Received",
                0, 0,
                amount,
                party["balance_due"],
                "Payment Received"
            ))

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("credit_transport"))

@app.route("/preview-duplicate-nozzle-entries")
def preview_duplicate_nozzle_entries():

    if not session.get("logged_in") or not is_admin():
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            ne.entry_date,
            nm.nozzle_name,
            nm.machine_no,
            COUNT(*) AS row_count
        FROM nozzle_entries ne
        LEFT JOIN nozzle_master nm ON ne.nozzle_id = nm.id
        GROUP BY ne.entry_date, ne.nozzle_id, nm.nozzle_name, nm.machine_no
        HAVING COUNT(*) > 1
        ORDER BY ne.entry_date DESC
    """)
    duplicate_groups = cur.fetchall()

    conn.close()

    return render_template(
        "duplicate_nozzle_entries.html",
        duplicate_groups=duplicate_groups
    )

@app.route("/cleanup-duplicate-nozzle-entries", methods=["POST"])
def cleanup_duplicate_nozzle_entries():

    if not session.get("logged_in") or not is_admin():
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        # for every (date, nozzle) that has more than one reading saved,
        # keep only the most recently saved row and delete the rest —
        # this is what was causing MS/HSD/CNG totals to show doubled
        cur.execute("""
            DELETE FROM nozzle_entries
            WHERE id IN (
                SELECT id FROM (
                    SELECT id,
                        ROW_NUMBER() OVER (
                            PARTITION BY entry_date, nozzle_id
                            ORDER BY id DESC
                        ) AS rn
                    FROM nozzle_entries
                ) ranked
                WHERE rn > 1
            )
            RETURNING id
        """)
        removed = cur.fetchall()

        log_activity(
            cur, "Nozzle Management", "Deleted",
            f"Cleaned up {len(removed)} duplicate nozzle reading rows (same date + nozzle saved more than once)"
        )

        conn.commit()

        flash(f"Removed {len(removed)} duplicate nozzle entries.")

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("settings"))

@app.route("/backup-database")
def backup_database():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))

    # the app only ever writes to Postgres (via DATABASE_URL) — the old
    # local sqlite file was never populated in production and Render's
    # filesystem is ephemeral anyway, so route this to the real export
    return redirect(url_for("full_system_export"))

# app.py
# DAILY CLOSING ROUTE

@app.route("/export-party-transport-excel")
def export_party_transport_excel():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    transporter_id = request.args.get("transporter_id")
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM credit_transporters
        WHERE id=%s
    """, (transporter_id,))
    party = cur.fetchone()

    party_name = party["party_name"] if party else "Transporter"

    cur.execute("""
        SELECT *
        FROM transport_entries
        WHERE transporter_id=%s
          AND entry_date >= %s
          AND entry_date <= %s
        ORDER BY entry_date ASC, sl_no ASC
    """, (transporter_id, from_date, to_date))

    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transport Entry Report"

    ws.append([
        "Date", "SL No", "Party Name", "Challan No", "Vehicle No",
        "Slip No", "HSD Qty", "Rate", "HSD Amount",
        "Diesel", "Final Amount"
    ])

    total_hsd = 0
    total_hsd_amount = 0
    total_cash = 0
    total_final = 0

    for r in rows:
        total_hsd += float(r["hsd_qty"] or 0)
        total_hsd_amount += float(r["hsd_amount"] or 0)
        total_cash += float(r["cash_taken"] or 0)
        total_final += float(r["total_amount"] or 0)

        ws.append([
            r["entry_date"],
            r["sl_no"],
            r["transporter_name"],
            r["challan_no"],
            r["vehicle_no"],
            r["slip_no"],
            r["hsd_qty"],
            r["rate"],
            r["hsd_amount"],
            r["cash_taken"],
            round(float(r["total_amount"] or 0))
        ])

    ws.append([])

    total_row = [
       "", "", "", "", "", "TOTAL",
       round(total_hsd, 2),
       "-",
       round(total_hsd_amount),
       round(total_cash),
       round(total_final)
]

    ws.append(total_row)
    total_row_number = ws.max_row

    discount_per_litre = round(float(party.get("discount") or 0), 2) if party else 0
    total_discount = round(discount_per_litre * total_hsd)

    if discount_per_litre > 0:
        net_payable = round(total_final) - total_discount

        ws.append([])
        ws.append(["", "", "", "", "", "", "", "", "", "Total Amount", round(total_final)])
        total_amt_row_number = ws.max_row

        ws.append(["", "", "", "", "", "", "", "", "", f"Discount (Rs.{discount_per_litre}/L x {total_hsd:.2f} L)", total_discount])
        discount_row_number = ws.max_row

        ws.append(["", "", "", "", "", "", "", "", "", "Net Payable", net_payable])
        net_row_number = ws.max_row

        for cell in ws[total_amt_row_number]:
            cell.font = Font(bold=True, color="07120C")
            cell.fill = PatternFill("solid", fgColor="F1F5F9")

        for cell in ws[discount_row_number]:
            cell.font = Font(bold=True, color="B91C1C")
            cell.fill = PatternFill("solid", fgColor="FEE2E2")

        for cell in ws[net_row_number]:
            cell.font = Font(bold=True, size=13, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="16A34A")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="07120C")

    for cell in ws[total_row_number]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")

    style_excel_sheet(ws)

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    filename = bill_filename(party_name, from_date, "xlsx")

    return send_file(
        file,
        as_attachment=True,
        download_name=filename
    )


@app.route("/export-party-transport-pdf")
def format_indian(n):
    """
    Formats a number using the Indian numbering system (lakhs/crores):
    rightmost 3 digits, then groups of 2 — e.g. 298050.80 -> "2,98,051"
    """
    n = int(round(float(n or 0)))
    negative = n < 0
    n = abs(n)
    s = str(n)

    if len(s) <= 3:
        result = s
    else:
        last3 = s[-3:]
        rest = s[:-3]
        parts = []
        while len(rest) > 2:
            parts.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        result = ",".join(parts) + "," + last3

    return ("-" if negative else "") + result


def export_party_transport_pdf():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    transporter_id = request.args.get("transporter_id")
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM credit_transporters
        WHERE id=%s
    """, (transporter_id,))
    party = cur.fetchone()

    party_name = party["party_name"] if party else "Transporter"

    cur.execute("SELECT * FROM settings WHERE id=1")
    biz = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM transport_entries
        WHERE transporter_id=%s
          AND entry_date >= %s
          AND entry_date <= %s
        ORDER BY entry_date ASC, sl_no ASC
    """, (transporter_id, from_date, to_date))

    rows = cur.fetchall()

    conn.close()

    station_name = (biz["station_name"] if biz and biz["station_name"] else "") or "SAI FUEL MART"
    station_address = biz["station_address"] if biz and biz["station_address"] else ""
    gstin = biz["gstin"] if biz and biz["gstin"] else ""
    phone_number = biz["phone_number"] if biz and biz["phone_number"] else ""

    file = BytesIO()

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e5e7eb"))
        canvas.line(20, 32, 822, 32)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawRightString(822, 18, f"Page {doc.page}")
        canvas.drawString(20, 18, f"{station_name} — Credit Transport Bill — Computer generated, no signature required")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        file,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    elements = []

    station_title_style = ParagraphStyle(
        "StationTitle", parent=styles["Title"],
        fontSize=20, leading=22, textColor=colors.HexColor("#07120C"),
        alignment=0, spaceAfter=0
    )
    station_sub_style = ParagraphStyle(
        "StationSub", parent=styles["Normal"],
        fontSize=8.5, leading=12, textColor=colors.HexColor("#475467")
    )
    bill_title_style = ParagraphStyle(
        "BillTitle", parent=styles["Normal"],
        fontSize=13, leading=16, textColor=colors.white,
        alignment=2, fontName="Helvetica-Bold"
    )
    bill_meta_style = ParagraphStyle(
        "BillMeta", parent=styles["Normal"],
        fontSize=8.5, leading=13, textColor=colors.white, alignment=2
    )
    label_style = ParagraphStyle(
        "Label", parent=styles["Normal"],
        fontSize=9, leading=13, textColor=colors.HexColor("#07120C")
    )

    # letterhead: station info on the left, bill meta panel on the right
    station_block = [
        Paragraph(station_name.upper(), station_title_style)
    ]
    if station_address:
        station_block.append(Paragraph(station_address, station_sub_style))
    meta_bits = []
    if gstin:
        meta_bits.append(f"GSTIN: {gstin}")
    if phone_number:
        meta_bits.append(f"Ph: {phone_number}")
    if meta_bits:
        station_block.append(Paragraph(" | ".join(meta_bits), station_sub_style))

    bill_block = [
        Paragraph("CREDIT TRANSPORT BILL", bill_title_style),
        Paragraph(f"Bill No: {bill_filename(party_name, from_date, '').rstrip('.').replace(' ', '-')}", bill_meta_style),
        Paragraph(f"Period: {from_date} to {to_date}", bill_meta_style),
        Paragraph(f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", bill_meta_style),
    ]

    letterhead = Table(
        [[station_block, bill_block]],
        colWidths=[520, 262]
    )
    letterhead.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#07120C")),
        ("LEFTPADDING", (1, 0), (1, 0), 16),
        ("RIGHTPADDING", (1, 0), (1, 0), 16),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
    ]))
    elements.append(letterhead)
    elements.append(Spacer(1, 4))

    # a thin green accent rule under the letterhead
    accent = Table([[""]], colWidths=[782], rowHeights=[4])
    accent.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#16a34a"))]))
    elements.append(accent)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"<b>Bill To:</b> {party_name}", label_style))
    if party and party["vehicle_no"]:
        elements.append(Paragraph(f"<b>Vehicle No:</b> {party['vehicle_no']}", label_style))
    if party and party["mobile"]:
        elements.append(Paragraph(f"<b>Contact:</b> {party['mobile']}", label_style))
    elements.append(Spacer(1, 12))

    data = [[
        "Date", "SL", "Challan", "Vehicle",
        "Slip", "HSD Qty (L)", "Rate", "HSD Amt",
        "Diesel", "Final Amt"
    ]]

    total_hsd = 0
    total_hsd_amount = 0
    total_cash = 0
    total_final = 0

    for r in rows:
        hsd_qty = float(r["hsd_qty"] or 0)
        rate = float(r["rate"] or 0)
        hsd_amount = float(r["hsd_amount"] or 0)
        cash = float(r["cash_taken"] or 0)
        final = float(r["total_amount"] or 0)

        total_hsd += hsd_qty
        total_hsd_amount += hsd_amount
        total_cash += cash
        total_final += final

        data.append([
            str(r["entry_date"]),
            str(r["sl_no"]),
            r["challan_no"] or "-",
            r["vehicle_no"] or "-",
            r["slip_no"] or "-",
            f"{hsd_qty:.2f}",
            f"{rate:.2f}",
            f"{hsd_amount:.2f}",
            f"{cash:.2f}",
            f"{final:.2f}"
        ])

    if not rows:
        data.append(["No entries in this date range", "", "", "", "", "", "", "", "", ""])

    data.append([
        "", "", "", "", "TOTAL",
        f"{total_hsd:.2f}",
        "-",
        f"{total_hsd_amount:.2f}",
        f"{total_cash:.2f}",
        f"{format_indian(total_final)}.00"
    ])

    total_row_index = len(data) - 1

    discount_per_litre = round(float(party.get("discount") or 0), 2) if party else 0
    total_discount = round(discount_per_litre * total_hsd, 2)

    if discount_per_litre > 0:
        net_payable = round(total_final - total_discount, 2)

        data.append([
            "", "", "", "", "", "", "", "", "Total Amount",
            f"{format_indian(total_final)}.00"
        ])
        data.append([
            "", "", "", "", "", "", "", "",
            f"Discount (Rs.{discount_per_litre}/L x {total_hsd:.2f}L)",
            f"{format_indian(total_discount)}.00"
        ])
        data.append([
            "", "", "", "", "", "", "", "", "Net Payable",
            f"{format_indian(net_payable)}.00"
        ])

    table = Table(
        data,
        colWidths=[62, 26, 75, 80, 60, 68, 55, 75, 75, 80],
        repeatRows=1
    )

    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#07120C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, total_row_index), (-1, total_row_index), "Helvetica-Bold"),

        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("LEADING", (0, 0), (-1, -1), 9),

        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("ALIGN", (5, 1), (-1, -1), "RIGHT"),

        ("BACKGROUND", (0, total_row_index), (-1, total_row_index), colors.HexColor("#DCFCE7")),
        ("TEXTCOLOR", (0, total_row_index), (-1, total_row_index), colors.HexColor("#07120C")),

        ("ROWBACKGROUNDS", (0, 1), (-1, total_row_index - 1), [
            colors.white,
            colors.HexColor("#F8FAFC")
        ]),
    ]

    if discount_per_litre > 0:
        total_amt_row_index = total_row_index + 1
        discount_row_index = total_row_index + 2
        net_row_index = len(data) - 1

        table_style.append(("SPAN", (0, total_amt_row_index), (7, total_amt_row_index)))
        table_style.append(("SPAN", (0, discount_row_index), (7, discount_row_index)))
        table_style.append(("SPAN", (0, net_row_index), (7, net_row_index)))

        table_style.append(("LINEABOVE", (0, total_amt_row_index), (-1, total_amt_row_index), 0.6, colors.HexColor("#94a3b8")))

        # Total Amount row — light grey, bold
        table_style.append(("BACKGROUND", (0, total_amt_row_index), (-1, total_amt_row_index), colors.HexColor("#F1F5F9")))
        table_style.append(("FONTNAME", (0, total_amt_row_index), (-1, total_amt_row_index), "Helvetica-Bold"))
        table_style.append(("TEXTCOLOR", (0, total_amt_row_index), (-1, total_amt_row_index), colors.HexColor("#07120C")))
        table_style.append(("FONTSIZE", (0, total_amt_row_index), (-1, total_amt_row_index), 8.5))

        # Discount row — red tint, bold
        table_style.append(("BACKGROUND", (0, discount_row_index), (-1, discount_row_index), colors.HexColor("#FEE2E2")))
        table_style.append(("FONTNAME", (0, discount_row_index), (-1, discount_row_index), "Helvetica-Bold"))
        table_style.append(("TEXTCOLOR", (0, discount_row_index), (-1, discount_row_index), colors.HexColor("#B91C1C")))
        table_style.append(("FONTSIZE", (0, discount_row_index), (-1, discount_row_index), 8.5))

        # Net Payable row — bold green, standout
        table_style.append(("BACKGROUND", (0, net_row_index), (-1, net_row_index), colors.HexColor("#16A34A")))
        table_style.append(("FONTNAME", (0, net_row_index), (-1, net_row_index), "Helvetica-Bold"))
        table_style.append(("TEXTCOLOR", (0, net_row_index), (-1, net_row_index), colors.white))
        table_style.append(("FONTSIZE", (0, net_row_index), (-1, net_row_index), 10))
        table_style.append(("TOPPADDING", (0, net_row_index), (-1, net_row_index), 9))
        table_style.append(("BOTTOMPADDING", (0, net_row_index), (-1, net_row_index), 9))

    table.setStyle(TableStyle(table_style))

    elements.append(table)
    elements.append(Spacer(1, 30))

    sig_data = [["", ""], ["Prepared By", "Authorized Signatory"]]
    sig = Table(sig_data, colWidths=[391, 391])
    sig.setStyle(TableStyle([
        ("LINEABOVE", (0, 1), (0, 1), 0.6, colors.HexColor("#94a3b8")),
        ("LINEABOVE", (1, 1), (1, 1), 0.6, colors.HexColor("#94a3b8")),
        ("TOPPADDING", (0, 1), (-1, 1), 4),
        ("FONTSIZE", (0, 1), (-1, 1), 8.5),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#667085")),
        ("ALIGN", (0, 1), (0, 1), "LEFT"),
        ("ALIGN", (1, 1), (1, 1), "RIGHT"),
    ]))
    elements.append(sig)

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)

    file.seek(0)

    filename = bill_filename(party_name, from_date, "pdf")

    return send_file(
        file,
        as_attachment=True,
        download_name=filename
    )

def reverse_daily_closing_for_date(cur, report_date):
    """
    Fully undo everything a Daily Closing save created for a given date:
    lube stock movements, transporter fuel/lube credit, and the matching
    ledger rows. Call this before re-saving an existing date (so nothing
    doubles up) and before deleting a closing (so no balance gets stuck).
    """

    # reverse lube stock and remove the lube-sale transactions this
    # closing created (both cash and credit sales)
    cur.execute("""
        SELECT id, product_id, qty
        FROM lube_transactions
        WHERE date=%s
          AND remarks IN ('Cash Lube Sale', 'Credit Lube Sale')
    """, (report_date,))

    for tx in cur.fetchall():
        qty = float(tx["qty"] or 0)
        if tx["product_id"] and qty:
            cur.execute("""
                UPDATE lube_stock
                SET sale_qty = COALESCE(sale_qty,0) - %s,
                    closing_stock = COALESCE(closing_stock,0) + %s
                WHERE id=%s
            """, (qty, qty, tx["product_id"]))

    cur.execute("""
        DELETE FROM lube_transactions
        WHERE date=%s
          AND remarks IN ('Cash Lube Sale', 'Credit Lube Sale')
    """, (report_date,))

    # reverse whatever fuel/lube credit this closing gave transporters
    cur.execute("""
        SELECT id, transporter_id, fuel_credit, lube_credit
        FROM transporter_ledger
        WHERE date=%s
          AND entry_type IN ('Fuel Credit', 'Lube Credit')
    """, (report_date,))

    for row in cur.fetchall():
        fuel_amt = round(float(row["fuel_credit"] or 0), 2)
        lube_amt = round(float(row["lube_credit"] or 0), 2)
        total_amt = round(fuel_amt + lube_amt, 2)

        if row["transporter_id"] and total_amt:
            cur.execute("""
                UPDATE credit_transporters
                SET fuel_credit = COALESCE(fuel_credit,0) - %s,
                    lube_credit = COALESCE(lube_credit,0) - %s,
                    credit_given = COALESCE(credit_given,0) - %s,
                    balance_due = COALESCE(balance_due,0) - %s
                WHERE id=%s
            """, (fuel_amt, lube_amt, total_amt, total_amt, row["transporter_id"]))

    cur.execute("""
        DELETE FROM transporter_ledger
        WHERE date=%s
          AND entry_type IN ('Fuel Credit', 'Lube Credit')
    """, (report_date,))


@app.route("/save-daily-closing", methods=["POST"])
def save_daily_closing():

    if not session.get("logged_in"):
        return jsonify({"status": "error", "message": "Not logged in"})

    data = request.get_json()

    if not data or not data.get("date"):
        return jsonify({"status": "error", "message": "Date missing"})

    conn = None

    try:
        conn = get_pg_conn()
        cur = conn.cursor()

        cur.execute("SELECT id FROM daily_closing WHERE date=%s", (data["date"],))
        existing = cur.fetchone()

        if existing:
            reverse_daily_closing_for_date(cur, data["date"])
            cur.execute("DELETE FROM daily_closing WHERE id=%s", (existing["id"],))
            cur.execute("DELETE FROM nozzle_entries WHERE entry_date=%s", (data["date"],))

        cur.execute("""
            INSERT INTO daily_closing (
                date, ms_litres, hsd_litres, cng_sale,
                total_fuel_sale, lube_sale, digital_collection,
                phonepe, card_swipe, hp_pay, hpcl_otp, upi_other,
                credit_given, transport_received, net_credit_due,
                total_expense, cash_in_hand
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data["date"],
            float(data.get("ms_litres", 0)),
            float(data.get("hsd_litres", 0)),
            float(data.get("cng_sale", 0)),
            float(data.get("total_fuel_sale", 0)),
            float(data.get("lube_sale", 0)),
            float(data.get("digital_collection", 0)),
            float(data.get("phonepe", 0)),
            float(data.get("card_swipe", 0)),
            float(data.get("hp_pay", 0)),
            float(data.get("hpcl_otp", 0)),
            float(data.get("upi_other", 0)),
            float(data.get("credit_given", 0)),
            float(data.get("transport_received", 0)),
            float(data.get("net_credit_due", 0)),
            float(data.get("total_expense", 0)),
            float(data.get("cash_in_hand", 0))
        ))

        # SAVE NOZZLES
        for item in data.get("nozzle_entries", []):

            nozzle_id = item.get("nozzle_id")

            if not nozzle_id:
                continue

            opening = float(item.get("opening", 0))
            closing = float(item.get("closing", 0))
            testing = float(item.get("testing", 0))

            sale = max(round(closing - opening - testing, 2), 0)

            cur.execute("""
                INSERT INTO nozzle_entries (
                    entry_date, nozzle_id, opening_reading,
                    closing_reading, testing_qty, total_sale, created_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                data["date"],
                nozzle_id,
                opening,
                closing,
                testing,
                sale,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))

        # SAVE LUBE CASH / CREDIT
        for item in data.get("lube_sales", []):

            qty = float(item.get("qty", 0))
            rate = float(item.get("rate", 0))
            amount = float(item.get("amount", 0))

            product_id = item.get("product_id")
            product_name = item.get("product_name", "")

            mode = item.get("mode", "Cash")
            transporter_id = item.get("transporter_id", "")
            transporter_name = item.get("transporter_name", "")

            if qty <= 0 or not product_id:
                continue

            cur.execute("""
                UPDATE lube_stock
                SET sale_qty = COALESCE(sale_qty,0) + %s,
                    closing_stock = COALESCE(closing_stock,0) - %s
                WHERE id=%s
            """, (qty, qty, product_id))

            cur.execute("""
                INSERT INTO lube_transactions (
                    date, product_id, product_name, transaction_type,
                    qty, rate, amount, remarks
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data["date"],
                product_id,
                product_name,
                "Sale",
                qty,
                rate,
                amount,
                f"{mode} Lube Sale"
            ))

            if mode == "Credit":

                if not transporter_id:
                    return jsonify({
                        "status": "error",
                        "message": f"Please select transporter for credit lube: {product_name}"
                    })

                cur.execute("""
                    UPDATE credit_transporters
                    SET lube_credit = COALESCE(lube_credit,0) + %s,
                        credit_given = COALESCE(credit_given,0) + %s,
                        balance_due = COALESCE(balance_due,0) + %s
                    WHERE id=%s
                """, (
                    amount,
                    amount,
                    amount,
                    transporter_id
                ))

                cur.execute("""
                    SELECT party_name, balance_due
                    FROM credit_transporters
                    WHERE id=%s
                """, (transporter_id,))

                tr = cur.fetchone()

                cur.execute("""
                    INSERT INTO transporter_ledger (
                        date, transporter_id, transporter_name,
                        entry_type, lube_credit, balance_after, remarks
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    data["date"],
                    transporter_id,
                    tr["party_name"] if tr else transporter_name,
                    "Lube Credit",
                    amount,
                    tr["balance_due"] if tr else amount,
                    product_name
                ))

        # SAVE FUEL CREDIT FROM DAILY CLOSING
        for item in data.get("credit_transport_sales", []):

            amount = float(item.get("amount", 0))
            transporter_id = item.get("transporter_id", "")

            if amount <= 0 or not transporter_id:
                continue

            cur.execute("""
                UPDATE credit_transporters
                SET fuel_credit = COALESCE(fuel_credit,0) + %s,
                    credit_given = COALESCE(credit_given,0) + %s,
                    balance_due = COALESCE(balance_due,0) + %s
                WHERE id=%s
            """, (
                amount,
                amount,
                amount,
                transporter_id
            ))

            cur.execute("""
                SELECT party_name, balance_due
                FROM credit_transporters
                WHERE id=%s
            """, (transporter_id,))

            tr = cur.fetchone()

            cur.execute("""
                INSERT INTO transporter_ledger (
                    date, transporter_id, transporter_name,
                    entry_type, fuel_credit, balance_after, remarks
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                data["date"],
                transporter_id,
                tr["party_name"] if tr else "",
                "Fuel Credit",
                amount,
                tr["balance_due"] if tr else amount,
                "Fuel Credit from Daily Closing"
            ))

        log_activity(
            cur, "Daily Closing",
            "Updated" if existing else "Created",
            f"Daily closing for {data['date']} saved (Fuel: {data.get('total_fuel_sale',0)}, Lube: {data.get('lube_sale',0)})"
        )

        conn.commit()

        return jsonify({
            "status": "success",
            "message": "Daily closing saved successfully"
        })

    except Exception as e:

        if conn:
            conn.rollback()

        return jsonify({
            "status": "error",
            "message": f"Error: {str(e)}"
        })

    finally:

        if conn:
            conn.close()

@app.route("/save-lube-product", methods=["POST"])
def save_lube_product():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    product_name = request.form.get("product_name")
    selling_rate = float(request.form.get("selling_rate") or 0)
    opening_stock = float(request.form.get("opening_stock") or 0)
    purchase_qty = float(request.form.get("purchase_qty") or 0)

    closing_stock = opening_stock + purchase_qty

    cur.execute("""

        INSERT INTO lube_stock (

            product_name,
            selling_rate,
            opening_stock,
            purchase_qty,
            sale_qty,
            closing_stock

        )

        VALUES (%s, %s, %s, %s, %s, %s)

    """, (

        product_name,
        selling_rate,
        opening_stock,
        purchase_qty,
        0,
        closing_stock

    ))

    log_activity(
        cur, "Lube Stock", "Created",
        f"Added product '{product_name}' (rate ₹{selling_rate}, opening stock {opening_stock})"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("lube_stock"))

@app.route("/daily-closing")
def daily_closing():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    selected_date = request.args.get(
        "date",
        datetime.now().strftime("%Y-%m-%d")
    )

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT *
    FROM lube_stock
    ORDER BY product_name ASC
""")

    lube_items = cur.fetchall()

    cur.execute("""
    SELECT *
    FROM credit_transporters
    ORDER BY party_name ASC
""")
    transporters = cur.fetchall()

    # SETTINGS

    cur.execute("""
        SELECT *
        FROM settings
        WHERE id=1
    """)

    settings = cur.fetchone()

    # FETCH ALL NOZZLES

    cur.execute("""

    SELECT

        nm.id,
        nm.nozzle_name,
        nm.machine_no,
        nm.fuel_type,

        COALESCE(

            ne.opening_reading,

            (

                SELECT prev.closing_reading

                FROM nozzle_entries prev

                WHERE prev.nozzle_id = nm.id
                AND prev.entry_date < %s

                ORDER BY prev.entry_date DESC,
                         prev.id DESC

                LIMIT 1

            ),

            0

        ) AS opening_reading,

        ne.closing_reading,
        ne.testing_qty,
        ne.total_sale

    FROM nozzle_master nm

    LEFT JOIN nozzle_entries ne

    ON nm.id = ne.nozzle_id
    AND ne.entry_date = %s

    ORDER BY
    nm.fuel_type,
    nm.nozzle_name

""", (selected_date, selected_date))

        

    all_nozzles = cur.fetchall()

    ms_nozzles = []
    hsd_nozzles = []
    cng_nozzles = []

    for n in all_nozzles:

        if n["fuel_type"] == "MS":
            ms_nozzles.append(n)

        elif n["fuel_type"] == "HSD":
            hsd_nozzles.append(n)

        elif n["fuel_type"] == "CNG":
            cng_nozzles.append(n)

    conn.close()

    return render_template(

        "daily_closing.html",

        settings=settings,
        transporters=transporters,
        ms_nozzles=ms_nozzles,
        hsd_nozzles=hsd_nozzles,
        cng_nozzles=cng_nozzles,
        lube_items=lube_items,
        today=selected_date
    )
@app.route("/print-daily-report/<date>")
def print_daily_report(date):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM daily_closing
        WHERE TO_CHAR(date, 'YYYY-MM-DD') = %s
        ORDER BY id DESC
        LIMIT 1
    """, (date,))
    closing = cur.fetchone()

    cur.execute("""
        SELECT
            nozzle_entries.*,
            nozzle_master.nozzle_name,
            nozzle_master.fuel_type,
            nozzle_master.machine_no
        FROM nozzle_entries
        LEFT JOIN nozzle_master
        ON nozzle_entries.nozzle_id = nozzle_master.id
        WHERE TO_CHAR(nozzle_entries.entry_date, 'YYYY-MM-DD') = %s
        ORDER BY nozzle_master.fuel_type, nozzle_master.nozzle_name
    """, (date,))
    nozzle_rows = cur.fetchall()

    cur.execute("""
        SELECT *
        FROM tank_level
        WHERE TO_CHAR(date, 'YYYY-MM-DD') = %s
        ORDER BY fuel_type ASC, id DESC
    """, (date,))
    tank_rows = cur.fetchall()

    conn.close()

    return render_template(
        "daily_report.html",
        closing=closing,
        nozzle_rows=nozzle_rows,
        tank_rows=tank_rows,
        report_date=date
    )


@app.route("/edit-transport-entry/<int:id>")
def edit_transport_entry(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM transport_entries WHERE id=%s",
        (id,)
    )

    entry = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM credit_transporters
        ORDER BY party_name
    """)

    transporters = cur.fetchall()

    conn.close()

    return render_template(
        "edit_transport_entry.html",
        entry=entry,
        transporters=transporters
    )

@app.route(
    "/update-transport-entry/<int:id>",
    methods=["POST"]
)
def update_transport_entry(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute(
            "SELECT * FROM transport_entries WHERE id=%s",
            (id,)
        )
        old_entry = cur.fetchone()

        if not old_entry:
            return redirect(url_for("credit_transport"))

        new_entry_date = request.form["entry_date"]
        new_transporter_id = request.form["transporter_id"]

        hsd_qty = float(request.form.get("hsd_qty") or 0)
        rate = float(request.form.get("rate") or 0)
        cash_taken = float(request.form.get("cash_taken") or 0)

        hsd_amount = round(hsd_qty * rate, 2)
        total_amount = round(hsd_amount + cash_taken, 2)

        challan_no = request.form.get("challan_no","")
        vehicle_no = request.form.get("vehicle_no","")
        slip_no = request.form.get("slip_no","")

        old_transporter_id = old_entry["transporter_id"]
        old_total_amount = round(float(old_entry["total_amount"] or 0), 2)
        old_entry_date = old_entry["entry_date"]

        cur.execute(
            "SELECT party_name FROM credit_transporters WHERE id=%s",
            (new_transporter_id,)
        )
        new_transporter_row = cur.fetchone()
        new_transporter_name = new_transporter_row["party_name"] if new_transporter_row else ""

        cur.execute("""
            UPDATE transport_entries
            SET
                entry_date=%s,
                transporter_id=%s,
                transporter_name=%s,
                challan_no=%s,
                vehicle_no=%s,
                slip_no=%s,
                hsd_qty=%s,
                rate=%s,
                hsd_amount=%s,
                cash_taken=%s,
                total_amount=%s
            WHERE id=%s
        """, (
            new_entry_date,
            new_transporter_id,
            new_transporter_name,
            challan_no,
            vehicle_no,
            slip_no,
            hsd_qty,
            rate,
            hsd_amount,
            cash_taken,
            total_amount,
            id
        ))

        if not old_transporter_id:
            # legacy entry that predates transporter linking — nothing was
            # ever credited for it, so just apply the full new amount now
            party = apply_transporter_credit_delta(
                cur, new_transporter_id, fuel_delta=total_amount
            )
            if party:
                cur.execute("""
                    INSERT INTO transporter_ledger(
                        date, transporter_id, transporter_name, entry_type,
                        fuel_credit, lube_credit, received_amount, balance_after, remarks
                    )
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,(
                    new_entry_date, new_transporter_id, new_transporter_name,
                    "Entry Updated",
                    total_amount, 0, 0, party["balance_due"],
                    f"Entry #{id} linked to a transporter for the first time"
                ))
        elif str(old_transporter_id) != str(new_transporter_id):
            # entry moved to a different transporter: fully reverse the old
            # party's credit and apply the full new amount to the new party
            old_party = apply_transporter_credit_delta(
                cur, old_transporter_id, fuel_delta=-old_total_amount
            )
            if old_party:
                cur.execute("""
                    INSERT INTO transporter_ledger(
                        date, transporter_id, transporter_name, entry_type,
                        fuel_credit, lube_credit, received_amount, balance_after, remarks
                    )
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,(
                    new_entry_date, old_transporter_id, old_party["party_name"],
                    "Entry Reassigned",
                    -old_total_amount, 0, 0, old_party["balance_due"],
                    f"Entry #{id} moved to another transporter"
                ))

            new_party = apply_transporter_credit_delta(
                cur, new_transporter_id, fuel_delta=total_amount
            )
            if new_party:
                cur.execute("""
                    INSERT INTO transporter_ledger(
                        date, transporter_id, transporter_name, entry_type,
                        fuel_credit, lube_credit, received_amount, balance_after, remarks
                    )
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,(
                    new_entry_date, new_transporter_id, new_transporter_name,
                    "Entry Reassigned",
                    total_amount, 0, 0, new_party["balance_due"],
                    f"Entry #{id} moved from another transporter"
                ))
        else:
            delta = round(total_amount - old_total_amount, 2)
            if delta:
                party = apply_transporter_credit_delta(
                    cur, new_transporter_id, fuel_delta=delta
                )
                if party:
                    cur.execute("""
                        INSERT INTO transporter_ledger(
                            date, transporter_id, transporter_name, entry_type,
                            fuel_credit, lube_credit, received_amount, balance_after, remarks
                        )
                        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,(
                        new_entry_date, new_transporter_id, new_transporter_name,
                        "Entry Updated",
                        delta, 0, 0, party["balance_due"],
                        f"Entry #{id} amount changed from Rs.{old_total_amount} to Rs.{total_amount}"
                    ))

        # keep SL No gap-free in both the old and new date groups
        if old_entry_date != new_entry_date:
            renumber_transport_entries(cur, old_entry_date)

        renumber_transport_entries(cur, new_entry_date)

        log_activity(
            cur, "Credit Transport", "Updated",
            f"Edited transport entry #{id} — ₹{old_total_amount} → ₹{total_amount}"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(
        url_for("credit_transport")
    )

@app.route("/edit-lube-transaction/<int:id>")
def edit_lube_transaction(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM lube_transactions
        WHERE id=%s
    """, (id,))
    tx = cur.fetchone()

    conn.close()

    return render_template(
        "edit_lube_transaction.html",
        tx=tx
    )

@app.route("/get-daily-closing/<date>")
def get_daily_closing(date):
    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM daily_closing WHERE date = %s ORDER BY id DESC LIMIT 1", (date,))
    row = cur.fetchone()

    cur.execute("""
        SELECT COALESCE(SUM(payment_amount), 0) AS total_received
        FROM transport_payments
        WHERE date <= %s
    """, (date,))
    transport_total = cur.fetchone()["total_received"]

    conn.close()

    if row:
        result = dict(row)
        result["transport_total"] = transport_total
        return jsonify(result)

    return jsonify({"transport_total": transport_total})

# =========================================
# EDIT DAILY CLOSING
# =========================================

@app.route("/save-salary-payment", methods=["POST"])
def save_salary_payment():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO salary_payments(
            payment_date,
            emp_id,
            employee_name,
            payment_mode,
            bank_account,
            amount,
            month_name,
            remarks
        )
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        request.form["payment_date"],
        request.form["emp_id"],
        request.form["employee_name"],
        request.form["payment_mode"],
        request.form.get("bank_account",""),
        float(request.form["amount"]),
        request.form["month_name"],
        request.form.get("remarks","")
    ))

    log_activity(
        cur, "Attendance", "Created",
        f"Salary payment of ₹{request.form['amount']} to {request.form['employee_name']} for {request.form['month_name']}"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("attendance"))


@app.route("/edit-attendance/<int:id>")
def edit_attendance(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM attendance
        WHERE id=%s
    """, (id,))
    att = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM staff_master
        ORDER BY staff_name ASC
    """)
    staff_list = cur.fetchall()

    conn.close()

    return render_template(
        "edit_attendance.html",
        att=att,
        staff_list=staff_list
    )

@app.route("/update-attendance/<int:id>", methods=["POST"])
def update_attendance(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        UPDATE attendance
        SET date=%s,
            staff_name=%s,
            attendance_status=%s
        WHERE id=%s
    """, (
        request.form["date"],
        request.form["staff_name"],
        request.form["attendance_status"],
        id
    ))

    log_activity(
        cur, "Attendance", "Updated",
        f"Attendance for {request.form['staff_name']} on {request.form['date']} set to {request.form['attendance_status']}"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("attendance"))

@app.route("/delete-attendance/<int:id>")
def delete_attendance(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT staff_name, date FROM attendance WHERE id=%s", (id,))
    att = cur.fetchone()

    cur.execute("""
        DELETE FROM attendance
        WHERE id=%s
    """, (id,))

    if att:
        log_activity(
            cur, "Attendance", "Deleted",
            f"Deleted attendance for {att['staff_name']} on {att['date']}"
        )

    conn.commit()
    conn.close()

    return redirect(url_for("attendance"))

@app.route("/edit-staff/<int:id>")
def edit_staff(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM staff_master
        WHERE id=%s
    """, (id,))

    staff = cur.fetchone()

    conn.close()

    return render_template(
        "edit_staff.html",
        staff=staff
    )
@app.route("/update-staff/<int:id>", methods=["POST"])
def update_staff(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT staff_name
        FROM staff_master
        WHERE id=%s
    """, (id,))

    old = cur.fetchone()

    old_name = old["staff_name"]

    cur.execute("""
        UPDATE staff_master
        SET
            emp_id=%s,
            staff_name=%s,
            role=%s,
            department=%s,
            joined_date=%s,
            bank_account=%s,
            shift=%s,
            status=%s
        WHERE id=%s
    """, (

        request.form.get("emp_id"),
        request.form.get("staff_name"),
        request.form.get("role"),
        request.form.get("department"),
        request.form.get("joined_date"),
        request.form.get("bank_account"),
        request.form.get("shift"),
        request.form.get("status"),
        id
    ))

    # UPDATE ATTENDANCE NAME ALSO

    cur.execute("""
        UPDATE attendance
        SET staff_name=%s
        WHERE staff_name=%s
    """, (
        request.form.get("staff_name"),
        old_name
    ))

    # UPDATE SALARY PAYMENT HISTORY TOO, so a rename doesn't
    # orphan their past payments under the old name/emp id
    cur.execute("""
        UPDATE salary_payments
        SET employee_name=%s,
            emp_id=%s
        WHERE employee_name=%s
    """, (
        request.form.get("staff_name"),
        request.form.get("emp_id"),
        old_name
    ))

    log_activity(
        cur, "Attendance", "Updated",
        f"Updated staff '{old_name}' → '{request.form.get('staff_name')}'"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("attendance"))

@app.route("/delete-staff/<int:id>")
def delete_staff(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT staff_name
            FROM staff_master
            WHERE id=%s
        """, (id,))

        staff = cur.fetchone()

        if staff:

            staff_name = staff["staff_name"]

            cur.execute("""
                DELETE FROM attendance
                WHERE staff_name=%s
            """, (staff_name,))

            cur.execute("""
                DELETE FROM salary_payments
                WHERE employee_name=%s
            """, (staff_name,))

            cur.execute("""
                DELETE FROM staff_master
                WHERE id=%s
            """, (id,))

            log_activity(
                cur, "Attendance", "Deleted",
                f"Deleted staff '{staff_name}' and their attendance/salary history"
            )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("attendance"))




@app.route("/edit-daily-closing/<int:id>")
def edit_daily_closing(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM daily_closing
        WHERE id=%s
    """, (id,))

    row = cur.fetchone()

    conn.close()

    return render_template(
        "edit_daily_closing.html",
        row=row
    )


# =========================================
# UPDATE DAILY CLOSING
# =========================================

@app.route("/update-daily-closing/<int:id>",
methods=["POST"])
def update_daily_closing(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""

        UPDATE daily_closing

        SET

            date=%s,

            ms_litres=%s,
            hsd_litres=%s,
            cng_sale=%s,

            total_fuel_sale=%s,

            lube_sale=%s,

            digital_collection=%s,

            total_expense=%s,

            cash_in_hand=%s

        WHERE id=%s

    """, (

        request.form["date"],

        request.form["ms_litres"],
        request.form["hsd_litres"],
        request.form["cng_sale"],

        request.form["total_fuel_sale"],

        request.form["lube_sale"],

        request.form["digital_collection"],

        request.form["total_expense"],

        request.form["cash_in_hand"],

        id

    ))

    log_activity(
        cur, "Daily Closing", "Updated",
        f"Edited daily closing for {request.form['date']} (fuel sale ₹{request.form['total_fuel_sale']})"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("reports"))



@app.route("/nozzle-management")
def nozzle_management():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM nozzle_master
        ORDER BY fuel_type ASC, nozzle_name ASC
    """)
    nozzle_master_rows = cur.fetchall()

    cur.execute("""
        SELECT
            nozzle_entries.*,
            nozzle_master.nozzle_name,
            nozzle_master.machine_no,
            nozzle_master.fuel_type
        FROM nozzle_entries
        LEFT JOIN nozzle_master
        ON nozzle_entries.nozzle_id = nozzle_master.id
        ORDER BY nozzle_entries.entry_date DESC,
                 nozzle_entries.id DESC
    """)
    nozzle_rows = cur.fetchall()

    conn.close()

    return render_template(
        "nozzle_management.html",
        nozzle_master_rows=nozzle_master_rows,
        nozzle_rows=nozzle_rows
    )


@app.route("/save-nozzle-master", methods=["POST"])
def save_nozzle_master():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    nozzle_name = request.form.get("nozzle_name", "").strip()
    machine_no = request.form.get("machine_no", "").strip()
    fuel_type = request.form.get("fuel_type", "").strip()
    status = request.form.get("status", "Active").strip()

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO nozzle_master (
            nozzle_name,
            machine_no,
            fuel_type,
            status
        )
        VALUES (%s, %s, %s, %s)
    """, (
        nozzle_name,
        machine_no,
        fuel_type,
        status
    ))

    log_activity(
        cur, "Nozzle Management", "Created",
        f"Added nozzle '{nozzle_name}' ({fuel_type}, machine {machine_no})"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("nozzle_management"))


@app.route("/edit-nozzle-master/<int:id>")
def edit_nozzle_master(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM nozzle_master
        WHERE id = %s
    """, (id,))
    nozzle = cur.fetchone()

    conn.close()

    return render_template(
        "edit_nozzle_master.html",
        nozzle=nozzle
    )


@app.route("/update-nozzle-master/<int:id>", methods=["POST"])
def update_nozzle_master(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    nozzle_name = request.form.get("nozzle_name", "").strip()
    machine_no = request.form.get("machine_no", "").strip()
    fuel_type = request.form.get("fuel_type", "").strip()
    status = request.form.get("status", "Active").strip()

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        UPDATE nozzle_master
        SET nozzle_name = %s,
            machine_no = %s,
            fuel_type = %s,
            status = %s
        WHERE id = %s
    """, (
        nozzle_name,
        machine_no,
        fuel_type,
        status,
        id
    ))

    log_activity(
        cur, "Nozzle Management", "Updated",
        f"Updated nozzle '{nozzle_name}' ({fuel_type}, machine {machine_no})"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("nozzle_management"))


@app.route("/delete-nozzle-master/<int:id>")
def delete_nozzle_master(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT nozzle_name FROM nozzle_master WHERE id=%s", (id,))
    nz = cur.fetchone()

    cur.execute("""
        DELETE FROM nozzle_entries
        WHERE nozzle_id = %s
    """, (id,))

    cur.execute("""
        DELETE FROM nozzle_master
        WHERE id = %s
    """, (id,))

    log_activity(
        cur, "Nozzle Management", "Deleted",
        f"Deleted nozzle '{nz['nozzle_name'] if nz else id}' and its readings"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("nozzle_management"))


@app.route("/save-nozzle-entry", methods=["POST"])
def save_nozzle_entry():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    entry_date = request.form.get("entry_date", "")
    nozzle_id = request.form.get("nozzle_id", "")

    opening_reading = float(request.form.get("opening_reading") or 0)
    closing_reading = float(request.form.get("closing_reading") or 0)
    testing_qty = float(request.form.get("testing_qty") or 0)

    total_sale = round(
      closing_reading
      - opening_reading
      - testing_qty,
    2
)

    if total_sale < 0:
     total_sale = 0

    conn = get_pg_conn()
    cur = conn.cursor()

    # a nozzle can only have ONE reading per day — check for an existing
    # entry first so re-saving corrects it instead of duplicating it
    # (duplicates were silently doubling every fuel-sale total downstream)
    cur.execute("""
        SELECT id
        FROM nozzle_entries
        WHERE entry_date=%s AND nozzle_id=%s
    """, (entry_date, nozzle_id))

    existing = cur.fetchone()

    if existing:

        cur.execute("""
            UPDATE nozzle_entries
            SET opening_reading=%s,
                closing_reading=%s,
                testing_qty=%s,
                total_sale=%s,
                created_at=%s
            WHERE id=%s
        """, (
            opening_reading,
            closing_reading,
            testing_qty,
            total_sale,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            existing["id"]
        ))

        log_activity(
            cur, "Nozzle Management", "Updated",
            f"Nozzle reading entry for {entry_date} (nozzle #{nozzle_id}, sale {total_sale} L) — corrected re-save"
        )

    else:

        cur.execute("""
            INSERT INTO nozzle_entries (
                entry_date,
                nozzle_id,
                opening_reading,
                closing_reading,
                testing_qty,
                total_sale,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            entry_date,
            nozzle_id,
            opening_reading,
            closing_reading,
            testing_qty,
            total_sale,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

        log_activity(
            cur, "Nozzle Management", "Created",
            f"Nozzle reading entry for {entry_date} (nozzle #{nozzle_id}, sale {total_sale} L)"
        )

    conn.commit()
    conn.close()

    return redirect(url_for("nozzle_management"))


@app.route("/edit-nozzle-entry/<int:id>")
def edit_nozzle_entry(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM nozzle_entries
        WHERE id = %s
    """, (id,))
    entry = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM nozzle_master
        ORDER BY fuel_type ASC, nozzle_name ASC
    """)
    nozzles = cur.fetchall()

    conn.close()

    return render_template(
        "edit_nozzle_entry.html",
        entry=entry,
        nozzles=nozzles
    )


@app.route("/update-nozzle-entry/<int:id>", methods=["POST"])
def update_nozzle_entry(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    entry_date = request.form.get("entry_date", "")
    nozzle_id = request.form.get("nozzle_id", "")

    opening_reading = float(request.form.get("opening_reading") or 0)
    closing_reading = float(request.form.get("closing_reading") or 0)
    testing_qty = float(request.form.get("testing_qty") or 0)

    total_sale = round(
        closing_reading - opening_reading - testing_qty,
        2
    )

    if total_sale < 0:
        total_sale = 0

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        UPDATE nozzle_entries
        SET entry_date = %s,
            nozzle_id = %s,
            opening_reading = %s,
            closing_reading = %s,
            testing_qty = %s,
            total_sale = %s,
            created_at = %s
        WHERE id = %s
    """, (
        entry_date,
        nozzle_id,
        opening_reading,
        closing_reading,
        testing_qty,
        total_sale,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        id
    ))

    log_activity(
        cur, "Nozzle Management", "Updated",
        f"Edited nozzle reading entry for {entry_date} (nozzle #{nozzle_id}, sale {total_sale} L)"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("nozzle_management"))


@app.route("/delete-nozzle-entry/<int:id>")
def delete_nozzle_entry(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM nozzle_entries
        WHERE id = %s
    """, (id,))

    log_activity(
        cur, "Nozzle Management", "Deleted",
        f"Deleted nozzle reading entry #{id}"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("nozzle_management"))


@app.route("/get-nozzle-sales/<date>")
def get_nozzle_sales(date):

    if not session.get("logged_in"):
        return jsonify([])

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""

        SELECT

            nozzle_entries.*,

            nozzle_master.nozzle_name,
            nozzle_master.machine_no,
            nozzle_master.fuel_type

        FROM nozzle_entries

        LEFT JOIN nozzle_master
        ON nozzle_entries.nozzle_id = nozzle_master.id

        WHERE TO_CHAR(nozzle_entries.entry_date, 'YYYY-MM-DD')=%s

        ORDER BY nozzle_master.fuel_type,
                 nozzle_master.nozzle_name

    """, (date,))

    rows = cur.fetchall()

    result = []

    for row in rows:

        result.append({

    "nozzle_name":
    row["nozzle_name"],

    "fuel_type":
    row["fuel_type"],

    "machine_no":
    row["machine_no"],

    "opening":
    row["opening_reading"],

    "closing":
    row["closing_reading"],

    "testing":
    row["testing_qty"],

    "sale":
    row["total_sale"]

})

    conn.close()

    return jsonify(result)


@app.route("/tank-level")
def tank_level():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM tank_level ORDER BY date DESC, id DESC")
    rows = cur.fetchall()

    cur.execute("""
        SELECT
            SUM(gain_qty) AS total_gain_qty,
            SUM(shortage_qty) AS total_shortage_qty,
            SUM(gain_amount) AS total_gain_amount,
            SUM(shortage_amount) AS total_shortage_amount
        FROM tank_level
        WHERE TO_CHAR(date, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')
    """)
    monthly = cur.fetchone()

    cur.execute("SELECT * FROM tank_level WHERE fuel_type='MS' ORDER BY id DESC LIMIT 1")
    ms_latest = cur.fetchone()

    cur.execute("SELECT * FROM tank_level WHERE fuel_type='HSD' ORDER BY id DESC LIMIT 1")
    hsd_latest = cur.fetchone()

    def days_remaining(fuel_type, latest_row):
        if not latest_row or not latest_row["current_stock"]:
            return None

        cur.execute("""
            SELECT AVG(sale_stock) AS avg_sale
            FROM (
                SELECT sale_stock
                FROM tank_level
                WHERE fuel_type=%s AND sale_stock > 0
                ORDER BY date DESC
                LIMIT 7
            ) recent
        """, (fuel_type,))
        avg_row = cur.fetchone()
        avg_sale = float(avg_row["avg_sale"] or 0) if avg_row else 0

        if avg_sale <= 0:
            return None

        return round(float(latest_row["current_stock"]) / avg_sale, 1)

    ms_days_remaining = days_remaining("MS", ms_latest)
    hsd_days_remaining = days_remaining("HSD", hsd_latest)

    conn.close()

    return render_template(
        "tank_level.html",
        rows=rows,
        monthly=monthly,
        ms_latest=ms_latest,
        hsd_latest=hsd_latest,
        ms_days_remaining=ms_days_remaining,
        hsd_days_remaining=hsd_days_remaining
    )


@app.route("/api/tank-status")
def api_tank_status():
    if not session.get("logged_in"):
        return jsonify({"status": "error", "message": "Not logged in"}), 401

    conn = get_pg_conn()
    cur = conn.cursor()

    def safe_row(row):
        if not row:
            return None
        out = {}
        for key, value in dict(row).items():
            if isinstance(value, (int, float)):
                out[key] = value
            elif value is None:
                out[key] = None
            else:
                try:
                    out[key] = float(value)
                except (TypeError, ValueError):
                    out[key] = str(value)
        return out

    cur.execute("SELECT * FROM tank_level WHERE fuel_type='MS' ORDER BY id DESC LIMIT 1")
    ms_latest = safe_row(cur.fetchone())

    cur.execute("SELECT * FROM tank_level WHERE fuel_type='HSD' ORDER BY id DESC LIMIT 1")
    hsd_latest = safe_row(cur.fetchone())

    conn.close()

    return jsonify({
        "status": "success",
        "ms_latest": ms_latest,
        "hsd_latest": hsd_latest,
        "server_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })


@app.route("/save-tank-level", methods=["POST"])
def save_tank_level():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    fuel_type = request.form["fuel_type"]
    opening_stock = float(request.form["opening_stock"] or 0)
    received_stock = float(request.form["received_stock"] or 0)
    own_tanker_stock = float(request.form["own_tanker_stock"] or 0)
    actual_dip = float(request.form["dip_stock"] or 0)

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT ms_litres, hsd_litres
        FROM daily_closing
        WHERE date = %s
        ORDER BY id DESC
        LIMIT 1
    """, (request.form["date"],))
    closing_data = cur.fetchone()

    sale_stock = 0
    if closing_data:
        sale_stock = float(closing_data["ms_litres"] or 0) if fuel_type == "MS" else float(closing_data["hsd_litres"] or 0)

    theoretical_stock = opening_stock + received_stock + own_tanker_stock - sale_stock
    difference = round(actual_dip - theoretical_stock, 2)

    cur.execute("SELECT ms_rate, hsd_rate FROM settings WHERE id = 1")
    settings_data = cur.fetchone()

    rate = float(settings_data["ms_rate"]) if fuel_type == "MS" else float(settings_data["hsd_rate"])

    gain_qty = difference if difference > 0 else 0
    shortage_qty = abs(difference) if difference < 0 else 0
    gain_amount = round(gain_qty * rate, 2)
    shortage_amount = round(shortage_qty * rate, 2)

    cur.execute("""
        INSERT INTO tank_level (
            date, fuel_type, opening_stock, received_stock, own_tanker_stock,
            sale_stock, gain_qty, shortage_qty, current_stock,
            gain_amount, shortage_amount, created_at
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        request.form["date"],
        fuel_type,
        opening_stock,
        received_stock,
        own_tanker_stock,
        sale_stock,
        gain_qty,
        shortage_qty,
        actual_dip,
        gain_amount,
        shortage_amount,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))

    log_activity(
        cur, "Tank Level", "Created",
        f"Tank level entry for {fuel_type} on {request.form['date']} (dip {actual_dip} L, diff {difference} L)"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("tank_level"))


@app.route("/api/tank-sync", methods=["POST", "OPTIONS"])
def api_tank_sync():
    """
    Receives live tank readings pushed from the HP Smart Connect ATG
    bookmarklet (running in the browser at the pump). Cross-origin by
    design — the ATG computer's browser calls this directly — so it
    needs its own CORS headers and a shared-secret check instead of a
    normal login session.
    """

    if request.method == "OPTIONS":
        resp = jsonify({"status": "ok"})
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp

    payload = request.get_json(force=True, silent=True) or {}

    expected_secret = os.environ.get("TANK_SYNC_SECRET", "")

    if not expected_secret or payload.get("secret") != expected_secret:
        resp = jsonify({"status": "error", "message": "Invalid or missing secret"})
        resp.headers["Access-Control-Allow-Origin"] = "*"
        return resp, 403

    tanks = payload.get("tanks", [])
    today = datetime.now().strftime("%Y-%m-%d")

    conn = get_pg_conn()
    cur = conn.cursor()

    synced = []

    try:
        cur.execute("""
            SELECT ms_litres, hsd_litres
            FROM daily_closing
            WHERE date = %s
            ORDER BY id DESC
            LIMIT 1
        """, (today,))
        closing_data = cur.fetchone()

        cur.execute("SELECT ms_rate, hsd_rate FROM settings WHERE id = 1")
        settings_data = cur.fetchone()

        for t in tanks:
            fuel_type = (t.get("fuel_type") or "").upper()
            if fuel_type not in ("MS", "HSD"):
                continue

            net_volume = round(float(t.get("net_volume") or 0), 2)
            today_receipt = round(float(t.get("today_receipt_ltr") or 0), 2)
            today_sale_atg = round(float(t.get("today_sale_ltr") or 0), 2)
            capacity = round(float(t.get("capacity") or 0), 2)
            water_ltr = round(float(t.get("water_ltr") or 0), 2)
            temperature_c = round(float(t.get("temperature_c") or 0), 2)
            dip_mm = round(float(t.get("dip_mm") or 0), 2)
            gross_volume = round(float(t.get("gross_volume") or 0), 2)
            water_dip_mm = round(float(t.get("water_dip_mm") or 0), 2)
            ullage_ltr = round(float(t.get("ullage_ltr") or 0), 2)
            decantation_status = (t.get("decantation_status") or "")[:50]
            density_status = (t.get("density_status") or "")[:50]
            density_at_15 = round(float(t.get("density_at_15") or 0), 2)
            density_kg_m3 = round(float(t.get("density_kg_m3") or 0), 2)
            den_float_height = round(float(t.get("den_float_height") or 0), 2)

            # work backwards from the live reading to get today's opening stock
            opening_stock = round(net_volume - today_receipt + today_sale_atg, 2)
            received_stock = today_receipt
            own_tanker_stock = 0
            actual_dip = net_volume

            if closing_data:
                sale_stock = float(closing_data["ms_litres"] or 0) if fuel_type == "MS" else float(closing_data["hsd_litres"] or 0)
            else:
                # no daily closing saved yet today — use the ATG's own running total instead
                sale_stock = today_sale_atg

            theoretical_stock = opening_stock + received_stock + own_tanker_stock - sale_stock
            difference = round(actual_dip - theoretical_stock, 2)

            rate = float(settings_data["ms_rate"] or 0) if fuel_type == "MS" else float(settings_data["hsd_rate"] or 0)

            gain_qty = difference if difference > 0 else 0
            shortage_qty = abs(difference) if difference < 0 else 0
            gain_amount = round(gain_qty * rate, 2)
            shortage_amount = round(shortage_qty * rate, 2)

            # replace today's reading for this fuel instead of piling up
            # duplicates if the bookmarklet is clicked more than once a day
            cur.execute("""
                DELETE FROM tank_level
                WHERE date=%s AND fuel_type=%s
            """, (today, fuel_type))

            cur.execute("""
                INSERT INTO tank_level (
                    date, fuel_type, opening_stock, received_stock, own_tanker_stock,
                    sale_stock, gain_qty, shortage_qty, current_stock,
                    gain_amount, shortage_amount, created_at,
                    capacity, water_ltr, temperature_c,
                    dip_mm, gross_volume, water_dip_mm, ullage_ltr,
                    decantation_status, density_status, density_at_15,
                    density_kg_m3, den_float_height
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                today, fuel_type, opening_stock, received_stock, own_tanker_stock,
                sale_stock, gain_qty, shortage_qty, actual_dip,
                gain_amount, shortage_amount,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                capacity, water_ltr, temperature_c,
                dip_mm, gross_volume, water_dip_mm, ullage_ltr,
                decantation_status, density_status, density_at_15,
                density_kg_m3, den_float_height
            ))

            log_activity(
                cur, "Tank Level", "Created",
                f"Auto-synced {fuel_type} from ATG — live dip {actual_dip}L (was {sale_stock}L sold today)"
            )

            synced.append(fuel_type)

        conn.commit()

    except Exception as e:
        conn.rollback()
        conn.close()
        resp = jsonify({"status": "error", "message": str(e)})
        resp.headers["Access-Control-Allow-Origin"] = "*"
        return resp, 500

    conn.close()

    resp = jsonify({"status": "success", "synced": synced, "date": today})
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp


@app.route("/edit-tank-level/<int:id>")
def edit_tank_level(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM tank_level WHERE id = %s", (id,))
    row = cur.fetchone()

    conn.close()

    return render_template("edit_tank_level.html", row=row)


@app.route("/update-tank-level/<int:id>", methods=["POST"])
def update_tank_level(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    opening_stock = float(request.form["opening_stock"] or 0)
    received_stock = float(request.form["received_stock"] or 0)
    own_tanker_stock = float(request.form["own_tanker_stock"] or 0)
    sale_stock = float(request.form["sale_stock"] or 0)
    actual_dip = float(request.form["current_stock"] or 0)

    theoretical_stock = opening_stock + received_stock + own_tanker_stock - sale_stock
    difference = round(actual_dip - theoretical_stock, 2)

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT ms_rate, hsd_rate FROM settings WHERE id = 1")
    settings_data = cur.fetchone()

    rate = float(settings_data["ms_rate"]) if request.form["fuel_type"] == "MS" else float(settings_data["hsd_rate"])

    gain_qty = difference if difference > 0 else 0
    shortage_qty = abs(difference) if difference < 0 else 0
    gain_amount = round(gain_qty * rate, 2)
    shortage_amount = round(shortage_qty * rate, 2)

    cur.execute("""
        UPDATE tank_level
        SET date=%s, fuel_type=%s, opening_stock=%s, received_stock=%s,
            own_tanker_stock=%s, sale_stock=%s, gain_qty=%s, shortage_qty=%s,
            current_stock=%s, gain_amount=%s, shortage_amount=%s, created_at=%s
        WHERE id=%s
    """, (
        request.form["date"],
        request.form["fuel_type"],
        opening_stock,
        received_stock,
        own_tanker_stock,
        sale_stock,
        gain_qty,
        shortage_qty,
        actual_dip,
        gain_amount,
        shortage_amount,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        id
    ))

    log_activity(
        cur, "Tank Level", "Updated",
        f"Edited tank level entry #{id} for {request.form['fuel_type']} on {request.form['date']}"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("tank_level"))


@app.route("/delete-tank-level/<int:id>")
def delete_tank_level(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("DELETE FROM tank_level WHERE id = %s", (id,))

    log_activity(
        cur, "Tank Level", "Deleted",
        f"Deleted tank level entry #{id}"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("tank_level"))


@app.route("/add-tank-entry")
def add_tank_entry():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM fuel_receipts
        ORDER BY id DESC
        LIMIT 200
    """)
    receipts = cur.fetchall()

    # pre-fill vehicle/carrier from the most recent receipt — these are
    # almost always the same fixed tanker/carrier for this pump
    cur.execute("""
        SELECT vehicle_no, carrier_no, carrier_name
        FROM fuel_receipts
        ORDER BY id DESC
        LIMIT 1
    """)
    last_receipt = cur.fetchone()

    cur.execute("""
        SELECT
            COUNT(*) AS receipt_count,
            COALESCE(SUM(total_ms_vol), 0) AS total_ms,
            COALESCE(SUM(total_hsd_vol), 0) AS total_hsd
        FROM fuel_receipts
        WHERE TO_CHAR(receipt_date::date, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')
    """)
    month_stats = cur.fetchone()

    cur.execute("""
        SELECT receipt_date
        FROM fuel_receipts
        ORDER BY id DESC
        LIMIT 1
    """)
    last_date_row = cur.fetchone()

    conn.close()

    return render_template(
        "add_tank_entry.html",
        receipts=receipts,
        last_receipt=last_receipt,
        month_stats=month_stats,
        last_receipt_date=last_date_row["receipt_date"] if last_date_row else None
    )


@app.route("/save-fuel-receipt", methods=["POST"])
def save_fuel_receipt():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        comp_data = []
        total_ms_vol = 0
        total_hsd_vol = 0

        for n in range(1, 6):
            fuel = request.form.get(f"comp{n}_fuel", "")
            dip = round(float(request.form.get(f"comp{n}_dip") or 0), 2)
            vol = round(float(request.form.get(f"comp{n}_vol") or 0), 2)

            comp_data.extend([fuel, dip, vol])

            if fuel == "MS":
                total_ms_vol += vol
            elif fuel == "HSD":
                total_hsd_vol += vol

        cur.execute("""
            INSERT INTO fuel_receipts (
                receipt_date, invoice_no, order_no, vehicle_no,
                carrier_no, carrier_name, po_no, po_date,
                water_checked, density, temperature_c,
                dip_before, dip_after,
                comp1_fuel, comp1_dip, comp1_vol,
                comp2_fuel, comp2_dip, comp2_vol,
                comp3_fuel, comp3_dip, comp3_vol,
                comp4_fuel, comp4_dip, comp4_vol,
                comp5_fuel, comp5_dip, comp5_vol,
                total_ms_vol, total_hsd_vol, created_at
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            request.form.get("receipt_date"),
            request.form.get("invoice_no", ""),
            request.form.get("order_no", ""),
            request.form.get("vehicle_no", ""),
            request.form.get("carrier_no", ""),
            request.form.get("carrier_name", ""),
            request.form.get("po_no", ""),
            request.form.get("po_date", ""),
            request.form.get("water_checked", "NIL"),
            round(float(request.form.get("density") or 0), 2),
            round(float(request.form.get("temperature_c") or 0), 2),
            round(float(request.form.get("dip_before") or 0), 2),
            round(float(request.form.get("dip_after") or 0), 2),
            *comp_data,
            round(total_ms_vol, 2),
            round(total_hsd_vol, 2),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

        new_id = cur.fetchone()["id"]

        log_activity(
            cur, "Tank Level", "Created",
            f"Fuel receipt #{new_id} logged — Invoice {request.form.get('invoice_no','')} "
            f"(MS {total_ms_vol}L, HSD {total_hsd_vol}L)"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("add_tank_entry"))


@app.route("/edit-fuel-receipt/<int:id>")
def edit_fuel_receipt(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM fuel_receipts WHERE id=%s", (id,))
    receipt = cur.fetchone()

    conn.close()

    if not receipt:
        flash("Fuel receipt not found.")
        return redirect(url_for("add_tank_entry"))

    return render_template("edit_fuel_receipt.html", r=receipt)


@app.route("/update-fuel-receipt/<int:id>", methods=["POST"])
def update_fuel_receipt(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        comp_data = []
        total_ms_vol = 0
        total_hsd_vol = 0

        for n in range(1, 6):
            fuel = request.form.get(f"comp{n}_fuel", "")
            dip = round(float(request.form.get(f"comp{n}_dip") or 0), 2)
            vol = round(float(request.form.get(f"comp{n}_vol") or 0), 2)

            comp_data.extend([fuel, dip, vol])

            if fuel == "MS":
                total_ms_vol += vol
            elif fuel == "HSD":
                total_hsd_vol += vol

        cur.execute("""
            UPDATE fuel_receipts
            SET receipt_date=%s, invoice_no=%s, order_no=%s, vehicle_no=%s,
                carrier_no=%s, carrier_name=%s, po_no=%s, po_date=%s,
                water_checked=%s, density=%s, temperature_c=%s,
                dip_before=%s, dip_after=%s,
                comp1_fuel=%s, comp1_dip=%s, comp1_vol=%s,
                comp2_fuel=%s, comp2_dip=%s, comp2_vol=%s,
                comp3_fuel=%s, comp3_dip=%s, comp3_vol=%s,
                comp4_fuel=%s, comp4_dip=%s, comp4_vol=%s,
                comp5_fuel=%s, comp5_dip=%s, comp5_vol=%s,
                total_ms_vol=%s, total_hsd_vol=%s
            WHERE id=%s
        """, (
            request.form.get("receipt_date"),
            request.form.get("invoice_no", ""),
            request.form.get("order_no", ""),
            request.form.get("vehicle_no", ""),
            request.form.get("carrier_no", ""),
            request.form.get("carrier_name", ""),
            request.form.get("po_no", ""),
            request.form.get("po_date", ""),
            request.form.get("water_checked", "NIL"),
            round(float(request.form.get("density") or 0), 2),
            round(float(request.form.get("temperature_c") or 0), 2),
            round(float(request.form.get("dip_before") or 0), 2),
            round(float(request.form.get("dip_after") or 0), 2),
            *comp_data,
            round(total_ms_vol, 2),
            round(total_hsd_vol, 2),
            id
        ))

        log_activity(
            cur, "Tank Level", "Updated",
            f"Edited fuel receipt #{id} — Invoice {request.form.get('invoice_no','')} "
            f"(MS {total_ms_vol}L, HSD {total_hsd_vol}L)"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("add_tank_entry"))


@app.route("/delete-fuel-receipt/<int:id>")
def delete_fuel_receipt(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("DELETE FROM fuel_receipts WHERE id = %s", (id,))

    log_activity(
        cur, "Tank Level", "Deleted",
        f"Deleted fuel receipt #{id}"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("add_tank_entry"))


def bill_filename(party_name, from_date, extension):
    """
    Build a short filename like 'SVT Bill 110626.xlsx' from a party
    name and the report's start date, instead of a long descriptive one.
    """
    try:
        date_code = datetime.strptime(str(from_date), "%Y-%m-%d").strftime("%d%m%y")
    except Exception:
        date_code = re.sub(r"[^0-9]", "", str(from_date))

    safe_party = re.sub(r'[\\/:*?"<>|]', "", party_name or "Transporter").strip()

    return f"{safe_party} Bill {date_code}.{extension}"


def renumber_transport_entries(cur, entry_date):
    """
    Re-sequence SL No for every transport entry on a given date so that
    after an insert/move/delete the numbers stay 1..N with no gaps.
    """
    cur.execute("""
        SELECT id
        FROM transport_entries
        WHERE entry_date=%s
        ORDER BY id ASC
    """, (entry_date,))

    rows = cur.fetchall()

    for idx, row in enumerate(rows, start=1):
        cur.execute(
            "UPDATE transport_entries SET sl_no=%s WHERE id=%s",
            (idx, row["id"])
        )


def apply_transporter_credit_delta(cur, transporter_id, fuel_delta=0, lube_delta=0):
    """
    Adjust a transporter's fuel_credit / lube_credit / credit_given / balance_due
    by the given deltas (can be negative to reverse an entry) and return the
    updated transporter row. This is the single place that keeps the
    Transporter Credit Summary in sync with every entry that touches it.
    """
    fuel_delta = round(fuel_delta or 0, 2)
    lube_delta = round(lube_delta or 0, 2)
    total_delta = round(fuel_delta + lube_delta, 2)

    cur.execute("""
        UPDATE credit_transporters
        SET
            fuel_credit = COALESCE(fuel_credit,0) + %s,
            lube_credit = COALESCE(lube_credit,0) + %s,
            credit_given = COALESCE(credit_given,0) + %s,
            balance_due = COALESCE(balance_due,0) + %s
        WHERE id=%s
        RETURNING *
    """, (
        fuel_delta,
        lube_delta,
        total_delta,
        total_delta,
        transporter_id
    ))

    return cur.fetchone()


@app.route("/credit-transport")
def credit_transport():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM credit_transporters
        ORDER BY party_name
    """)
    transporters = cur.fetchall()

    cur.execute("""
        SELECT *
        FROM transport_entries
        ORDER BY entry_date DESC,id DESC
    """)
    entries = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*) as total
        FROM credit_transporters
    """)
    total_transporters = cur.fetchone()["total"]

    cur.execute("""
        SELECT COUNT(*) as total
        FROM transport_entries
        WHERE entry_date = CURRENT_DATE
    """)
    today_entries = cur.fetchone()["total"]

    cur.execute("""
SELECT COALESCE(SUM(credit_given), 0) AS total_credit
FROM credit_transporters
""")

    row = cur.fetchone()

    total_credit = row["total_credit"] if row else 0

    cur.execute("""
SELECT hsd_rate
FROM settings
LIMIT 1
""")

    row = cur.fetchone()

    hsd_rate = row["hsd_rate"] if row else 0

    conn.close()

    return render_template(
        "credit_transport.html",
        transporters=transporters,
        entries=entries,
        total_transporters=total_transporters,
        today_entries=today_entries,
        hsd_rate=hsd_rate,
        total_credit=total_credit
    )

@app.route("/add-transporter", methods=["POST"])
def add_transporter():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    party_name = request.form.get("party_name","").strip()

    if not party_name:
        flash("Party name is required.")
        return redirect(url_for("credit_transport"))

    opening_balance = round(float(request.form.get("opening_balance",0) or 0), 2)

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO credit_transporters(
                party_name,
                mobile,
                vehicle_no,
                opening_balance,
                balance_due,
                status
            )
            VALUES(%s,%s,%s,%s,%s,%s)
            RETURNING id, party_name
        """,(
            party_name,
            request.form.get("mobile","").strip(),
            request.form.get("vehicle_no","").strip(),
            opening_balance,
            opening_balance,
            request.form.get("status","Active")
        ))

        new_transporter = cur.fetchone()

        if opening_balance:
            cur.execute("""
                INSERT INTO transporter_ledger(
                    date, transporter_id, transporter_name, entry_type,
                    fuel_credit, lube_credit, received_amount, balance_after, remarks
                )
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,(
                datetime.now().strftime("%Y-%m-%d"),
                new_transporter["id"],
                new_transporter["party_name"],
                "Opening Balance",
                0, 0, 0,
                opening_balance,
                "Opening balance set when transporter was added"
            ))

        log_activity(
            cur, "Credit Transport", "Created",
            f"Added transporter '{party_name}' (opening balance ₹{opening_balance})"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("credit_transport"))

@app.route("/save-transport-entry", methods=["POST"])
def save_transport_entry():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    transporter_id = request.form.get("transporter_id")

    if not transporter_id:
        flash("Please select a transporter.")
        return redirect(url_for("credit_transport"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        entry_date = request.form["entry_date"]

        cur.execute("""
            SELECT COUNT(*) as total
            FROM transport_entries
            WHERE entry_date=%s
        """,(entry_date,))

        sl_no = cur.fetchone()["total"] + 1

        qty = float(request.form.get("hsd_qty") or 0)
        rate = float(request.form.get("rate") or 0)
        cash_taken = float(request.form.get("cash_taken") or 0)

        hsd_amount = round(qty * rate, 2)
        total_amount = round(hsd_amount + cash_taken, 2)

        transporter_name = request.form.get("transporter_name","")
        challan_no = request.form.get("challan_no","")
        vehicle_no = request.form.get("vehicle_no","")
        slip_no = request.form.get("slip_no","")

        cur.execute("""
            INSERT INTO transport_entries(
                entry_date,
                sl_no,
                transporter_id,
                transporter_name,
                challan_no,
                vehicle_no,
                slip_no,
                hsd_qty,
                rate,
                hsd_amount,
                cash_taken,
                total_amount
            )
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """,(
            entry_date,
            sl_no,
            transporter_id,
            transporter_name,
            challan_no,
            vehicle_no,
            slip_no,
            qty,
            rate,
            hsd_amount,
            cash_taken,
            total_amount
        ))

        new_entry_id = cur.fetchone()["id"]

        # keep the Transporter Credit Summary (fuel_credit / credit_given /
        # balance_due) perfectly in sync with this HSD credit entry
        party = apply_transporter_credit_delta(
            cur, transporter_id, fuel_delta=total_amount
        )

        cur.execute("""
            INSERT INTO transporter_ledger(
                date, transporter_id, transporter_name, entry_type,
                fuel_credit, lube_credit, received_amount, balance_after, remarks
            )
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,(
            entry_date,
            transporter_id,
            party["party_name"] if party else transporter_name,
            "Transport Entry",
            total_amount, 0, 0,
            party["balance_due"] if party else total_amount,
            f"Entry #{new_entry_id} | Challan {challan_no} | Slip {slip_no} | {qty} L @ Rs.{rate}"
        ))

        log_activity(
            cur, "Credit Transport", "Created",
            f"Transport entry #{new_entry_id} for {transporter_name} — ₹{total_amount} on {entry_date}"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("credit_transport"))

@app.route("/save-transport-payment", methods=["POST"])
def save_transport_payment():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    transporter_id = request.form.get("transporter_id")

    if not transporter_id:
        flash("Please select a transporter.")
        return redirect(url_for("credit_transport"))

    amount = round(float(request.form.get("payment_amount") or 0), 2)

    date = request.form["date"]

    payment_type = request.form.get("payment_type","Fuel")

    if amount <= 0:
        flash("Enter a valid amount received.")
        return redirect(url_for("credit_transport"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE credit_transporters
            SET
                payment_received = COALESCE(payment_received,0) + %s,
                balance_due = COALESCE(balance_due,0) - %s
            WHERE id=%s
            RETURNING *
        """,(
            amount,
            amount,
            transporter_id
        ))

        party = cur.fetchone()

        if party:
            cur.execute("""
                INSERT INTO transporter_ledger(
                    date,
                    transporter_id,
                    transporter_name,
                    entry_type,
                    fuel_credit,
                    lube_credit,
                    received_amount,
                    balance_after,
                    remarks
                )
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,(
                date,
                transporter_id,
                party["party_name"],
                f"{payment_type} Payment Received",
                0, 0,
                amount,
                party["balance_due"],
                "Payment Received"
            ))

            log_activity(
                cur, "Credit Transport", "Created",
                f"Received ₹{amount} ({payment_type}) from {party['party_name']} on {date}"
            )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("credit_transport"))

@app.route("/transporter-history/<int:id>")
def transporter_history(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM credit_transporters
        WHERE id=%s
    """,(id,))

    transporter = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM transporter_ledger
        WHERE transporter_id=%s
        ORDER BY id DESC
    """,(id,))

    history = cur.fetchall()

    conn.close()

    return render_template(
        "transporter_history.html",
        transporter=transporter,
        history=history
    )

@app.route("/edit-transporter-payment/<int:id>")
def edit_transporter_payment(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM transporter_ledger WHERE id=%s",
        (id,)
    )
    payment = cur.fetchone()

    conn.close()

    if not payment or not payment["received_amount"]:
        flash("Only 'Amount Received' entries can be edited here.")
        return redirect(url_for("credit_transport"))

    return render_template(
        "edit_transporter_payment.html",
        payment=payment
    )

@app.route("/update-transporter-payment/<int:id>", methods=["POST"])
def update_transporter_payment(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()
    transporter_id = None

    try:
        cur.execute(
            "SELECT * FROM transporter_ledger WHERE id=%s",
            (id,)
        )
        old_payment = cur.fetchone()

        if not old_payment or not old_payment["received_amount"]:
            flash("Only 'Amount Received' entries can be edited here.")
            return redirect(url_for("credit_transport"))

        transporter_id = old_payment["transporter_id"]
        old_amount = round(float(old_payment["received_amount"] or 0), 2)

        new_amount = round(float(request.form.get("received_amount") or 0), 2)
        new_date = request.form.get("date") or old_payment["date"]
        payment_type = request.form.get("payment_type", "Fuel")

        if new_amount <= 0:
            flash("Enter a valid amount received.")
            return redirect(url_for("transporter_history", id=transporter_id))

        # only the difference needs to move the balance
        delta = round(new_amount - old_amount, 2)

        cur.execute("""
            UPDATE credit_transporters
            SET
                payment_received = COALESCE(payment_received,0) + %s,
                balance_due = COALESCE(balance_due,0) - %s
            WHERE id=%s
            RETURNING balance_due
        """, (delta, delta, transporter_id))

        party = cur.fetchone()

        cur.execute("""
            UPDATE transporter_ledger
            SET
                date=%s,
                entry_type=%s,
                received_amount=%s,
                balance_after=%s,
                remarks=%s
            WHERE id=%s
        """, (
            new_date,
            f"{payment_type} Payment Received",
            new_amount,
            party["balance_due"] if party else old_payment["balance_after"],
            "Payment Received (edited)",
            id
        ))

        log_activity(
            cur, "Credit Transport", "Updated",
            f"Edited payment #{id}: ₹{old_amount} → ₹{new_amount}"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("transporter_history", id=transporter_id))

@app.route("/delete-transporter-payment/<int:id>")
def delete_transporter_payment(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()
    transporter_id = None

    try:
        cur.execute(
            "SELECT * FROM transporter_ledger WHERE id=%s",
            (id,)
        )
        payment = cur.fetchone()

        if not payment or not payment["received_amount"]:
            flash("Only 'Amount Received' entries can be deleted here.")
            return redirect(url_for("credit_transport"))

        transporter_id = payment["transporter_id"]
        amount = round(float(payment["received_amount"] or 0), 2)

        # reverse the payment so the balance goes right back to
        # what it was before this receipt was ever recorded
        cur.execute("""
            UPDATE credit_transporters
            SET
                payment_received = COALESCE(payment_received,0) - %s,
                balance_due = COALESCE(balance_due,0) + %s
            WHERE id=%s
        """, (amount, amount, transporter_id))

        cur.execute(
            "DELETE FROM transporter_ledger WHERE id=%s",
            (id,)
        )

        log_activity(
            cur, "Credit Transport", "Deleted",
            f"Deleted payment #{id} of ₹{amount} for {payment['transporter_name']}"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("transporter_history", id=transporter_id))

@app.route("/edit-transporter/<int:id>")
def edit_transporter(id):

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM credit_transporters
        WHERE id=%s
    """,(id,))

    transporter = cur.fetchone()

    conn.close()

    return render_template(
        "edit_transporter.html",
        transporter=transporter
    )

@app.route("/update-transporter/<int:id>",methods=["POST"])
def update_transporter(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    party_name = request.form.get("party_name","").strip()

    if not party_name:
        flash("Party name is required.")
        return redirect(url_for("credit_transport"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute(
            "SELECT opening_balance FROM credit_transporters WHERE id=%s",
            (id,)
        )
        existing = cur.fetchone()
        old_opening_balance = round(float(existing["opening_balance"] or 0), 2) if existing else 0

        new_opening_balance = round(float(request.form.get("opening_balance", 0) or 0), 2)
        opening_delta = round(new_opening_balance - old_opening_balance, 2)

        cur.execute("""
            UPDATE credit_transporters
            SET
            party_name=%s,
            mobile=%s,
            vehicle_no=%s,
            status=%s,
            opening_balance=%s,
            balance_due = COALESCE(balance_due,0) + %s
            WHERE id=%s
            RETURNING balance_due
        """,(
            party_name,
            request.form.get("mobile","").strip(),
            request.form.get("vehicle_no","").strip(),
            request.form.get("status","Active"),
            new_opening_balance,
            opening_delta,
            id
        ))
        updated = cur.fetchone()

        # keep transport_entries / ledger display names in sync with a rename
        cur.execute("""
            UPDATE transport_entries
            SET transporter_name=%s
            WHERE transporter_id=%s
        """, (party_name, id))

        cur.execute("""
            UPDATE transporter_ledger
            SET transporter_name=%s
            WHERE transporter_id=%s
        """, (party_name, id))

        if opening_delta:
            cur.execute("""
                INSERT INTO transporter_ledger(
                    date, transporter_id, transporter_name, entry_type,
                    fuel_credit, lube_credit, received_amount, balance_after, remarks
                )
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                datetime.now().strftime("%Y-%m-%d"),
                id, party_name, "Opening Balance Adjusted",
                0, 0, 0,
                updated["balance_due"] if updated else None,
                f"Opening balance changed from Rs.{old_opening_balance} to Rs.{new_opening_balance}"
            ))

        log_activity(
            cur, "Credit Transport", "Updated",
            f"Updated transporter details for '{party_name}'"
            + (f" — opening balance Rs.{old_opening_balance} to Rs.{new_opening_balance}" if opening_delta else "")
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("credit_transport"))

@app.route("/delete-daily-closing/<int:id>")
def delete_daily_closing(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute(
            "SELECT date FROM daily_closing WHERE id=%s",
            (id,)
        )

        row = cur.fetchone()

        if row:

            report_date = row["date"]

            # undo the stock/credit/ledger effects before removing the
            # closing itself, so nothing is left stuck on a transporter
            reverse_daily_closing_for_date(cur, report_date)

            cur.execute(
                "DELETE FROM daily_closing WHERE id=%s",
                (id,)
            )

            cur.execute(
                "DELETE FROM nozzle_entries WHERE entry_date=%s",
                (report_date,)
            )

            log_activity(
                cur, "Daily Closing", "Deleted",
                f"Deleted daily closing for {report_date} (credit/stock reversed)"
            )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("reports"))

@app.route("/delete-transporter/<int:id>")
def delete_transporter(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("SELECT party_name FROM credit_transporters WHERE id=%s", (id,))
        tp = cur.fetchone()

        # find every date group that will lose a row so SL No can be
        # re-sequenced after the cascading delete below
        cur.execute("""
            SELECT DISTINCT entry_date
            FROM transport_entries
            WHERE transporter_id=%s
        """, (id,))
        affected_dates = [row["entry_date"] for row in cur.fetchall()]

        cur.execute("DELETE FROM transport_entries WHERE transporter_id=%s", (id,))
        cur.execute("DELETE FROM transporter_ledger WHERE transporter_id=%s", (id,))
        cur.execute("DELETE FROM credit_transporters WHERE id=%s", (id,))

        for entry_date in affected_dates:
            renumber_transport_entries(cur, entry_date)

        log_activity(
            cur, "Credit Transport", "Deleted",
            f"Deleted transporter '{tp['party_name'] if tp else id}' and all their entries/ledger"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("credit_transport"))

@app.route("/delete-transport-entry/<int:id>")
def delete_transport_entry(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute(
            "SELECT * FROM transport_entries WHERE id=%s",
            (id,)
        )
        entry = cur.fetchone()

        if entry:
            cur.execute(
                "DELETE FROM transport_entries WHERE id=%s",
                (id,)
            )

            total_amount = round(float(entry["total_amount"] or 0), 2)
            transporter_id = entry["transporter_id"]

            # reverse the credit this entry had added, so the balance
            # never drifts out of sync after a delete
            if transporter_id and total_amount:
                party = apply_transporter_credit_delta(
                    cur, transporter_id, fuel_delta=-total_amount
                )
                if party:
                    cur.execute("""
                        INSERT INTO transporter_ledger(
                            date, transporter_id, transporter_name, entry_type,
                            fuel_credit, lube_credit, received_amount, balance_after, remarks
                        )
                        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,(
                        datetime.now().strftime("%Y-%m-%d"),
                        transporter_id,
                        party["party_name"],
                        "Entry Deleted",
                        -total_amount, 0, 0,
                        party["balance_due"],
                        f"Deleted entry #{id} dated {entry['entry_date']}"
                    ))

            # close the gap this delete left in the SL No sequence
            renumber_transport_entries(cur, entry["entry_date"])

            log_activity(
                cur, "Credit Transport", "Deleted",
                f"Deleted transport entry #{id} — ₹{total_amount} for {entry['transporter_name']}"
            )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("credit_transport"))


@app.route("/digital-collection")
def digital_collection():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")

    conn = get_pg_conn()
    cur = conn.cursor()

    if from_date and to_date:
        date_filter_sql = "WHERE date >= %s AND date <= %s"
        date_params = (from_date, to_date)
        range_label = f"{from_date} to {to_date}"
    else:
        # default view: current month, same as before
        date_filter_sql = "WHERE TO_CHAR(date, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')"
        date_params = ()
        range_label = "This Month"

    cur.execute(f"""
        SELECT
            SUM(digital_collection) AS total_digital,
            SUM(phonepe) AS total_phonepe,
            SUM(card_swipe) AS total_card,
            SUM(hp_pay) AS total_hp_pay,
            SUM(hpcl_otp) AS total_hpcl_otp,
            SUM(upi_other) AS total_upi,
            SUM(transport_received) AS total_transport
        FROM daily_closing
        {date_filter_sql}
    """, date_params)
    monthly = cur.fetchone()

    cur.execute(f"""
        SELECT date, digital_collection, phonepe, card_swipe,
               hp_pay, hpcl_otp, upi_other, transport_received
        FROM daily_closing
        {date_filter_sql}
        ORDER BY date DESC
        LIMIT 500
    """, date_params)
    rows = cur.fetchall()

    conn.close()

    return render_template(
        "digital_collection.html",
        rows=rows,
        monthly=monthly,
        from_date=from_date,
        to_date=to_date,
        range_label=range_label
    )


@app.route("/reports")
def reports():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    fuel_type = request.args.get("fuel_type", "")
    search = request.args.get("search", "")

    # DAILY CLOSING
    
    # DAILY CLOSING WITH NOZZLE-WISE HOVER DATA
    daily_query = """
     SELECT

        dc.*,

        COALESCE(nz.ms1, 0) AS ms1,
        COALESCE(nz.ms2, 0) AS ms2,
        COALESCE(nz.ms3, 0) AS ms3,
        COALESCE(nz.hsd1, 0) AS hsd1,
        COALESCE(nz.hsd2, 0) AS hsd2,
        COALESCE(nz.hsd3, 0) AS hsd3,
        COALESCE(nz.cng1, 0) AS cng1,
        COALESCE(nz.cng2, 0) AS cng2

    FROM daily_closing dc

    LEFT JOIN (
        SELECT
            TO_CHAR(ne.entry_date, 'YYYY-MM-DD') AS entry_day,
            SUM(CASE WHEN nm.nozzle_name = 'MS1' THEN ne.total_sale ELSE 0 END) AS ms1,
            SUM(CASE WHEN nm.nozzle_name = 'MS2' THEN ne.total_sale ELSE 0 END) AS ms2,
            SUM(CASE WHEN nm.nozzle_name = 'MS3' THEN ne.total_sale ELSE 0 END) AS ms3,
            SUM(CASE WHEN nm.nozzle_name = 'HSD1' THEN ne.total_sale ELSE 0 END) AS hsd1,
            SUM(CASE WHEN nm.nozzle_name = 'HSD2' THEN ne.total_sale ELSE 0 END) AS hsd2,
            SUM(CASE WHEN nm.nozzle_name = 'HSD3' THEN ne.total_sale ELSE 0 END) AS hsd3,
            SUM(CASE WHEN nm.nozzle_name = 'CNG1' THEN ne.total_sale ELSE 0 END) AS cng1,
            SUM(CASE WHEN nm.nozzle_name = 'CNG2' THEN ne.total_sale ELSE 0 END) AS cng2
        FROM nozzle_entries ne
        JOIN nozzle_master nm ON ne.nozzle_id = nm.id
        GROUP BY TO_CHAR(ne.entry_date, 'YYYY-MM-DD')
    ) nz ON nz.entry_day = TO_CHAR(dc.date, 'YYYY-MM-DD')

    WHERE 1=1
"""

    daily_params = []

    if from_date:
        daily_query += " AND dc.date >= %s"
        daily_params.append(from_date)

    if to_date:
        daily_query += " AND dc.date <= %s"
        daily_params.append(to_date)

    if search:
        daily_query += " AND TO_CHAR(dc.date, 'YYYY-MM-DD') ILIKE %s"
        daily_params.append(f"%{search}%")

    daily_query += " ORDER BY dc.date DESC, dc.id DESC"

    cur.execute(daily_query, daily_params)
    daily_rows = cur.fetchall()

    # NOZZLE REPORT
    nozzle_query = """
        SELECT
            nozzle_entries.*,
            nozzle_master.nozzle_name,
            nozzle_master.machine_no,
            nozzle_master.fuel_type
        FROM nozzle_entries
        LEFT JOIN nozzle_master
        ON nozzle_entries.nozzle_id = nozzle_master.id
        WHERE 1=1
    """
    nozzle_params = []

    if from_date:
        nozzle_query += " AND nozzle_entries.entry_date >= %s"
        nozzle_params.append(from_date)

    if to_date:
        nozzle_query += " AND nozzle_entries.entry_date <= %s"
        nozzle_params.append(to_date)

    if fuel_type:
        nozzle_query += " AND nozzle_master.fuel_type = %s"
        nozzle_params.append(fuel_type)

    if search:
        nozzle_query += """
            AND (
                nozzle_master.nozzle_name ILIKE %s
                OR nozzle_master.machine_no ILIKE %s
            )
        """
        nozzle_params.extend([f"%{search}%", f"%{search}%"])

    nozzle_query += """
        ORDER BY nozzle_entries.entry_date DESC, nozzle_entries.id DESC
    """

    cur.execute(nozzle_query, nozzle_params)
    nozzle_rows = cur.fetchall()

    # TANK REPORT
    tank_query = """
        SELECT *
        FROM tank_level
        WHERE 1=1
    """
    tank_params = []

    if from_date:
        tank_query += " AND date >= %s"
        tank_params.append(from_date)

    if to_date:
        tank_query += " AND date <= %s"
        tank_params.append(to_date)

    if fuel_type:
        tank_query += " AND fuel_type = %s"
        tank_params.append(fuel_type)

    tank_query += " ORDER BY date DESC, id DESC"

    cur.execute(tank_query, tank_params)
    tank_rows = cur.fetchall()

    # LUBE REPORT
    lube_query = """
        SELECT *
        FROM lube_stock
        WHERE 1=1
    """
    lube_params = []

    if search:
        lube_query += " AND product_name ILIKE %s"
        lube_params.append(f"%{search}%")

    lube_query += " ORDER BY product_name ASC"

    cur.execute(lube_query, lube_params)
    lube_rows = cur.fetchall()

    # CREDIT TRANSPORT REPORT
    transport_query = """
        SELECT *
        FROM credit_transporters
        WHERE 1=1
    """
    transport_params = []

    if search:
        transport_query += " AND party_name ILIKE %s"
        transport_params.append(f"%{search}%")

    transport_query += " ORDER BY balance_due DESC, party_name ASC"

    cur.execute(transport_query, transport_params)
    transporter_rows = cur.fetchall()

    # TRANSPORT ENTRY REPORT
    transport_entry_query = """
        SELECT *
        FROM transport_entries
        WHERE 1=1
    """
    transport_entry_params = []

    if from_date:
        transport_entry_query += " AND entry_date >= %s"
        transport_entry_params.append(from_date)

    if to_date:
        transport_entry_query += " AND entry_date <= %s"
        transport_entry_params.append(to_date)

    if search:
        transport_entry_query += """
            AND (
                transporter_name ILIKE %s
                OR challan_no ILIKE %s
                OR vehicle_no ILIKE %s
                OR slip_no ILIKE %s
            )
        """
        transport_entry_params.extend([
            f"%{search}%",
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ])

    transport_entry_query += " ORDER BY entry_date DESC, sl_no DESC"

    cur.execute(transport_entry_query, transport_entry_params)
    transport_entry_rows = cur.fetchall()

    transport_total_hsd = sum(float(r["hsd_qty"] or 0) for r in transport_entry_rows)
    transport_total_amount = sum(float(r["total_amount"] or 0) for r in transport_entry_rows)

    conn.close()

    return render_template(
        "reports.html",
        daily_rows=daily_rows,
        nozzle_rows=nozzle_rows,
        tank_rows=tank_rows,
        lube_rows=lube_rows,
        transporter_rows=transporter_rows,
        transport_entry_rows=transport_entry_rows,
        transport_total_hsd=transport_total_hsd,
        transport_total_amount=transport_total_amount
    )


@app.route("/export-full-backup")
def export_full_backup():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))

    conn = get_pg_conn()
    cur = conn.cursor()
    wb = Workbook()

    sheets = [
        ("Daily Closing", "SELECT * FROM daily_closing ORDER BY date DESC", [
            "Date", "MS Litres", "HSD Litres", "CNG Sale", "Fuel Sale", "Lube Sale", "Digital Collection",
            "PhonePe", "Card Swipe", "HP Pay", "HPCL OTP", "UPI Other", "Credit Given",
            "Transport Received", "Net Credit Due", "Expense", "Cash In Hand"
        ], lambda r: [
            r["date"], r["ms_litres"], r["hsd_litres"], r["cng_sale"], r["total_fuel_sale"], r["lube_sale"],
            r["digital_collection"], r["phonepe"], r["card_swipe"], r["hp_pay"], r["hpcl_otp"],
            r["upi_other"], r["credit_given"], r["transport_received"], r["net_credit_due"],
            r["total_expense"], r["cash_in_hand"]
        ]),

        ("Tank Level", "SELECT * FROM tank_level ORDER BY date DESC", [
            "Date", "Fuel", "Opening", "Received", "Own Tanker", "Sale", "Current",
            "Gain Qty", "Shortage Qty", "Gain Amount", "Shortage Amount", "Updated"
        ], lambda r: [
            r["date"], r["fuel_type"], r["opening_stock"], r["received_stock"], r["own_tanker_stock"],
            r["sale_stock"], r["current_stock"], r["gain_qty"], r["shortage_qty"],
            r["gain_amount"], r["shortage_amount"], r["created_at"]
        ]),

        ("Nozzle Master", "SELECT * FROM nozzle_master ORDER BY fuel_type, nozzle_name", [
            "Nozzle", "Machine No", "Fuel", "Status"
        ], lambda r: [
            r["nozzle_name"], r["machine_no"], r["fuel_type"], r["status"]
        ]),

        ("Nozzle Entries", """
            SELECT nozzle_entries.*, nozzle_master.nozzle_name, nozzle_master.fuel_type, nozzle_master.machine_no
            FROM nozzle_entries
            LEFT JOIN nozzle_master ON nozzle_entries.nozzle_id = nozzle_master.id
            ORDER BY nozzle_entries.entry_date DESC
        """, [
            "Date", "Nozzle", "Machine", "Fuel", "Opening", "Closing", "Testing", "Sale", "Updated"
        ], lambda r: [
            r["entry_date"], r["nozzle_name"], r["machine_no"], r["fuel_type"],
            r["opening_reading"], r["closing_reading"], r["testing_qty"], r["total_sale"], r["created_at"]
        ]),

        ("Lube Stock", "SELECT * FROM lube_stock ORDER BY product_name ASC", [
            "Product", "Rate", "Opening", "Purchase", "Sale", "Closing"
        ], lambda r: [
            r["product_name"], r["selling_rate"], r["opening_stock"], r["purchase_qty"],
            r["sale_qty"], r["closing_stock"]
        ]),

        ("Transporters", "SELECT * FROM credit_transporters ORDER BY party_name ASC", [
            "Party", "Opening", "Credit Given", "Payment Received", "Balance Due", "Status"
        ], lambda r: [
            r["party_name"], r["opening_balance"], r["credit_given"], r["payment_received"],
            r["balance_due"], r["status"]
        ]),

        ("Transport Entries", "SELECT * FROM transport_entries ORDER BY entry_date DESC, sl_no DESC", [
            "SL No", "Date", "Transport Name", "Challan No", "Vehicle No", "Slip No",
            "HSD Qty", "Rate", "HSD Amount", "Cash Taken", "Total Amount", "Updated"
        ], lambda r: [
            r["sl_no"], r["entry_date"], r["transporter_name"], r["challan_no"], r["vehicle_no"],
            r["slip_no"], r["hsd_qty"], r["rate"], r["hsd_amount"], r["cash_taken"],
            r["total_amount"], r["created_at"]
        ]),

        ("Attendance", "SELECT * FROM attendance ORDER BY date DESC", [
            "Date", "Staff Name", "Attendance"
        ], lambda r: [
            r["date"], r["staff_name"], r["attendance_status"]
        ])
    ]

    first = True

    for title, query, headers, row_func in sheets:
        ws = wb.active if first else wb.create_sheet(title)
        ws.title = title
        first = False

        ws.append(headers)

        cur.execute(query)
        rows = cur.fetchall()

        for row in rows:
            ws.append(row_func(row))

        style_excel_sheet(ws)

    conn.close()

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(
        file,
        as_attachment=True,
        download_name="Sai_Fuel_Mart_Full_Backup.xlsx"
    )


@app.route("/export-nozzle-report")
def export_nozzle_report():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))

    # export code

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            nozzle_entries.*,
            nozzle_master.nozzle_name,
            nozzle_master.fuel_type,
            nozzle_master.machine_no
        FROM nozzle_entries
        LEFT JOIN nozzle_master
        ON nozzle_entries.nozzle_id = nozzle_master.id
        ORDER BY nozzle_entries.entry_date DESC
    """)
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Nozzle Entries"

    ws.append([
        "Date", "Nozzle", "Machine", "Fuel",
        "Opening", "Closing", "Testing", "Sale", "Updated"
    ])

    for row in rows:
        ws.append([
            row["entry_date"], row["nozzle_name"], row["machine_no"],
            row["fuel_type"], row["opening_reading"], row["closing_reading"],
            row["testing_qty"], row["total_sale"], row["created_at"]
        ])

    style_excel_sheet(ws)

    conn.close()

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file, as_attachment=True, download_name="Nozzle_Report.xlsx")


@app.route("/export-reports")
def export_reports():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))

    # export code

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM daily_closing ORDER BY date DESC")
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Closing"

    ws.append([
        "Date", "MS Litres", "HSD Litres", "CNG Sale", "Fuel Sale", "Lube Sale",
        "Digital", "PhonePe", "Card", "HP Pay", "HPCL OTP", "UPI Other",
        "Credit Given", "Transport Received", "Net Credit Due", "Expense", "Cash"
    ])

    for row in rows:
        ws.append([
            row["date"], row["ms_litres"], row["hsd_litres"], row["cng_sale"], row["total_fuel_sale"],
            row["lube_sale"], row["digital_collection"], row["phonepe"], row["card_swipe"],
            row["hp_pay"], row["hpcl_otp"], row["upi_other"], row["credit_given"],
            row["transport_received"], row["net_credit_due"], row["total_expense"],
            row["cash_in_hand"]
        ])

    style_excel_sheet(ws)
    conn.close()

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file, as_attachment=True, download_name="Daily_Closing_Report.xlsx")


@app.route("/export-tank-report")
def export_tank_report():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM tank_level ORDER BY date DESC")
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tank Level"

    ws.append([
        "Date", "Fuel", "Opening", "Received", "Own Tanker", "Sale",
        "Current", "Gain Qty", "Shortage Qty", "Gain Amount", "Shortage Amount", "Updated"
    ])

    for row in rows:
        ws.append([
            row["date"], row["fuel_type"], row["opening_stock"], row["received_stock"],
            row["own_tanker_stock"], row["sale_stock"], row["current_stock"],
            row["gain_qty"], row["shortage_qty"], row["gain_amount"],
            row["shortage_amount"], row["created_at"]
        ])

    style_excel_sheet(ws)
    conn.close()

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file, as_attachment=True, download_name="Tank_Level_Report.xlsx")


@app.route("/export-lube-report")
def export_lube_report():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM lube_stock ORDER BY closing_stock ASC")
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lube Stock"

    ws.append(["Product", "Rate", "Opening", "Purchase", "Sale", "Closing"])

    for row in rows:
        ws.append([
            row["product_name"], row["selling_rate"], row["opening_stock"],
            row["purchase_qty"], row["sale_qty"], row["closing_stock"]
        ])

    style_excel_sheet(ws)
    conn.close()

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file, as_attachment=True, download_name="Lube_Stock_Report.xlsx")


@app.route("/export-transport-report")
def export_transport_report():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))
    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM credit_transporters ORDER BY balance_due DESC")
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transport Credit"

    ws.append(["Party", "Opening", "Credit Given", "Payment Received", "Balance Due", "Status"])

    for row in rows:
        ws.append([
            row["party_name"], row["opening_balance"], row["credit_given"],
            row["payment_received"], row["balance_due"], row["status"]
        ])

    style_excel_sheet(ws)
    conn.close()

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file, as_attachment=True, download_name="Transport_Credit_Report.xlsx")



@app.route("/export-transport-entry-report")
def export_transport_entry_report():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM transport_entries ORDER BY entry_date DESC, sl_no DESC")
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transport Entries"

    ws.append([
        "SL No", "Date", "Transport Name", "Challan No", "Vehicle No",
        "Slip No", "HSD Qty", "Rate", "HSD Amount", "Cash Taken",
        "Total Amount", "Updated"
    ])

    for row in rows:
        ws.append([
            row["sl_no"], row["entry_date"], row["transporter_name"], row["challan_no"],
            row["vehicle_no"], row["slip_no"], row["hsd_qty"], row["rate"],
            row["hsd_amount"], row["cash_taken"], row["total_amount"], row["created_at"]
        ])

    style_excel_sheet(ws)
    conn.close()

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file, as_attachment=True, download_name="Transport_Entries.xlsx")


@app.route("/delete-report/<int:id>")
def delete_report(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT date FROM daily_closing WHERE id=%s", (id,))
    row = cur.fetchone()

    if row:
        # this route duplicates /delete-daily-closing — reverse the same
        # way so it can't leave stuck transporter balances either
        reverse_daily_closing_for_date(cur, row["date"])

    cur.execute("DELETE FROM daily_closing WHERE id = %s", (id,))

    if row:
        log_activity(
            cur, "Daily Closing", "Deleted",
            f"Deleted daily closing for {row['date']} (via legacy route, credit/stock reversed)"
        )

    conn.commit()
    conn.close()

    return redirect(url_for("reports"))


@app.route("/edit-report/<int:id>")
def edit_report(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM daily_closing WHERE id = %s", (id,))
    row = cur.fetchone()
    conn.close()

    return render_template("edit_report.html", row=row)


@app.route("/update-report/<int:id>", methods=["POST"])
def update_report(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        UPDATE daily_closing
        SET date=%s, ms_litres=%s, hsd_litres=%s, cng_sale=%s, total_fuel_sale=%s,
            lube_sale=%s, digital_collection=%s, phonepe=%s, card_swipe=%s,
            hp_pay=%s, hpcl_otp=%s, upi_other=%s, credit_given=%s,
            transport_received=%s, net_credit_due=%s, total_expense=%s, cash_in_hand=%s
        WHERE id=%s
    """, (
        request.form["date"], request.form["ms_litres"], request.form["hsd_litres"],
        request.form.get("cng_sale",0), request.form["total_fuel_sale"], request.form["lube_sale"],
        request.form.get("digital_collection", 0), request.form.get("phonepe", 0),
        request.form.get("card_swipe", 0), request.form.get("hp_pay", 0),
        request.form.get("hpcl_otp", 0), request.form.get("upi_other", 0),
        request.form["credit_given"], request.form["transport_received"],
        request.form["net_credit_due"], request.form["total_expense"],
        request.form["cash_in_hand"], id
    ))

    log_activity(
        cur, "Daily Closing", "Updated",
        f"Edited daily closing for {request.form['date']} (via legacy route)"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("reports"))


@app.route("/settings")
def settings():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM settings WHERE id=1")
    settings_data = cur.fetchone()

    transporter_count = 0
    if is_admin():
        cur.execute("SELECT COUNT(*) AS total FROM credit_transporters")
        transporter_count = cur.fetchone()["total"]

    conn.close()

    return render_template(
        "settings.html",
        settings=settings_data,
        transporter_count=transporter_count
    )


@app.route("/settings-business-info")
def settings_business_info():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM settings WHERE id=1")
    settings_data = cur.fetchone()
    conn.close()

    return render_template("settings_business_info.html", settings=settings_data)


@app.route("/settings-fuel-rates")
def settings_fuel_rates():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM settings WHERE id=1")
    settings_data = cur.fetchone()
    conn.close()

    return render_template("settings_fuel_rates.html", settings=settings_data)


@app.route("/settings-transporter-discounts")
def settings_transporter_discounts():
    if not session.get("logged_in") or not is_admin():
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id, party_name, discount
            FROM credit_transporters
            ORDER BY party_name ASC
        """)
        transporters = cur.fetchall()
    except Exception:
        # self-heal: a startup migration hiccup may have skipped adding
        # this column — add it now and retry once instead of crashing
        conn.rollback()
        cur.execute("""
            ALTER TABLE credit_transporters ADD COLUMN IF NOT EXISTS discount REAL DEFAULT 0
        """)
        conn.commit()
        cur.execute("""
            SELECT id, party_name, discount
            FROM credit_transporters
            ORDER BY party_name ASC
        """)
        transporters = cur.fetchall()

    conn.close()

    return render_template(
        "settings_transporter_discounts.html",
        transporters=transporters
    )

# =========================================
# SAVE SETTINGS
# =========================================

@app.route("/save-business-info", methods=["POST"])
def save_business_info():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    station_name = request.form.get("station_name", "").strip()
    station_address = request.form.get("station_address", "").strip()
    gstin = request.form.get("gstin", "").strip()
    phone_number = request.form.get("phone_number", "").strip()

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("SELECT id FROM settings LIMIT 1")
        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE settings
                SET station_name=%s, station_address=%s, gstin=%s, phone_number=%s
                WHERE id=%s
            """, (station_name, station_address, gstin, phone_number, existing["id"]))
        else:
            cur.execute("""
                INSERT INTO settings (station_name, station_address, gstin, phone_number)
                VALUES (%s, %s, %s, %s)
            """, (station_name, station_address, gstin, phone_number))

        log_activity(
            cur, "Settings", "Updated",
            f"Business info updated — {station_name}"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("settings_business_info"))


@app.route("/save-fuel-rates", methods=["POST"])
def save_fuel_rates():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    ms_rate = float(request.form.get("ms_rate", 0) or 0)
    hsd_rate = float(request.form.get("hsd_rate", 0) or 0)
    cng_rate = float(request.form.get("cng_rate", 0) or 0)

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("SELECT id FROM settings LIMIT 1")
        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE settings
                SET ms_rate=%s, hsd_rate=%s, cng_rate=%s
                WHERE id=%s
            """, (ms_rate, hsd_rate, cng_rate, existing["id"]))
        else:
            cur.execute("""
                INSERT INTO settings (ms_rate, hsd_rate, cng_rate)
                VALUES (%s, %s, %s)
            """, (ms_rate, hsd_rate, cng_rate))

        log_activity(
            cur, "Settings", "Updated",
            f"Fuel rates updated — MS ₹{ms_rate}, HSD ₹{hsd_rate}, CNG ₹{cng_rate}"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("settings_fuel_rates"))


@app.route("/save-transporter-discounts", methods=["POST"])
def save_transporter_discounts():

    if not session.get("logged_in") or not is_admin():
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            ALTER TABLE credit_transporters ADD COLUMN IF NOT EXISTS discount REAL DEFAULT 0
        """)
        cur.execute("SELECT id, party_name, discount FROM credit_transporters")
        all_transporters = cur.fetchall()

        changed = []

        for t in all_transporters:
            field_name = f"discount_{t['id']}"

            if field_name not in request.form:
                continue

            new_discount = round(float(request.form.get(field_name) or 0), 2)
            old_discount = round(float(t["discount"] or 0), 2)

            if new_discount != old_discount:
                cur.execute("""
                    UPDATE credit_transporters
                    SET discount=%s
                    WHERE id=%s
                """, (new_discount, t["id"]))
                changed.append(f"{t['party_name']}: Rs.{old_discount} → Rs.{new_discount}")

        if changed:
            log_activity(
                cur, "Settings", "Updated",
                "Transporter discounts updated — " + "; ".join(changed)
            )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("settings_transporter_discounts"))


@app.route("/lube-stock")
def lube_stock():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM lube_stock
        ORDER BY product_name ASC
    """)
    lube_items = cur.fetchall()

    cur.execute("""
        SELECT *
        FROM lube_transactions
        ORDER BY date DESC, id DESC
    """)
    lube_transactions = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*) AS total_products,
               COALESCE(SUM(closing_stock),0) AS total_stock,
               COALESCE(SUM(sale_qty),0) AS total_sold
        FROM lube_stock
    """)
    summary = cur.fetchone()

    product_data = {}

    for item in lube_items:
        product_id = item["id"]

        cur.execute("""
            SELECT *
            FROM lube_transactions
            WHERE product_id=%s
            ORDER BY date DESC, id DESC
        """, (product_id,))
        transactions = cur.fetchall()

        total_purchase = sum(float(t["qty"] or 0) for t in transactions if t["transaction_type"] == "Purchase")
        total_sale = sum(float(t["qty"] or 0) for t in transactions if t["transaction_type"] == "Sale")

        product_data[product_id] = {
            "transactions": transactions,
            "total_purchase": total_purchase,
            "total_sale": total_sale
        }

    conn.close()

    return render_template(
        "lube_stock.html",
        lube_items=lube_items,
        lube_transactions=lube_transactions,
        product_data=product_data,
        summary=summary
    )

@app.route("/save-lube-transaction", methods=["POST"])
def save_lube_transaction():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    date = request.form.get("date")
    product_id = request.form.get("product_id")
    transaction_type = request.form.get("transaction_type")
    qty = float(request.form.get("qty") or 0)
    rate = float(request.form.get("rate") or 0)
    remarks = request.form.get("remarks", "")

    amount = round(qty * rate, 2)

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT product_name
            FROM lube_stock
            WHERE id=%s
        """, (product_id,))
        product = cur.fetchone()

        if not product:
            return redirect(url_for("lube_stock"))

        product_name = product["product_name"]

        cur.execute("""
            INSERT INTO lube_transactions (
                date, product_id, product_name, transaction_type,
                qty, rate, amount, remarks
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            date, product_id, product_name, transaction_type,
            qty, rate, amount, remarks
        ))

        if transaction_type == "Purchase":
            cur.execute("""
                UPDATE lube_stock
                SET purchase_qty = purchase_qty + %s,
                    closing_stock = closing_stock + %s
                WHERE id=%s
            """, (qty, qty, product_id))

        elif transaction_type == "Sale":
            cur.execute("""
                UPDATE lube_stock
                SET sale_qty = sale_qty + %s,
                    closing_stock = closing_stock - %s
                WHERE id=%s
            """, (qty, qty, product_id))

        log_activity(
            cur, "Lube Stock", "Created",
            f"{transaction_type} of {qty} {product_name} @ ₹{rate} on {date}"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("lube_stock"))

@app.route("/delete-lube-transaction/<int:id>")
def delete_lube_transaction(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM lube_transactions
            WHERE id=%s
        """, (id,))
        tx = cur.fetchone()

        if tx:
            product_id = tx["product_id"]
            qty = float(tx["qty"] or 0)

            if tx["transaction_type"] == "Purchase":
                cur.execute("""
                    UPDATE lube_stock
                    SET purchase_qty = purchase_qty - %s,
                        closing_stock = closing_stock - %s
                    WHERE id=%s
                """, (qty, qty, product_id))

            elif tx["transaction_type"] == "Sale":
                cur.execute("""
                    UPDATE lube_stock
                    SET sale_qty = sale_qty - %s,
                        closing_stock = closing_stock + %s
                    WHERE id=%s
                """, (qty, qty, product_id))

            cur.execute("""
                DELETE FROM lube_transactions
                WHERE id=%s
            """, (id,))

            log_activity(
                cur, "Lube Stock", "Deleted",
                f"Deleted {tx['transaction_type']} of {qty} {tx['product_name']} on {tx['date']}"
            )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("lube_stock"))

@app.route("/edit-lube/<int:id>")
def edit_lube(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM lube_stock
        WHERE id=%s
    """, (id,))

    item = cur.fetchone()

    conn.close()

    return render_template(
        "edit_lube.html",
        item=item
    )

@app.route("/update-lube-transaction/<int:id>", methods=["POST"])
def update_lube_transaction(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM lube_transactions
            WHERE id=%s
        """, (id,))
        old = cur.fetchone()

        if not old:
            return redirect(url_for("lube_stock"))

        old_qty = float(old["qty"] or 0)
        old_type = old["transaction_type"]
        product_id = old["product_id"]

        # reverse old stock effect
        if old_type == "Purchase":
            cur.execute("""
                UPDATE lube_stock
                SET purchase_qty = purchase_qty - %s,
                    closing_stock = closing_stock - %s
                WHERE id=%s
            """, (old_qty, old_qty, product_id))

        elif old_type == "Sale":
            cur.execute("""
                UPDATE lube_stock
                SET sale_qty = sale_qty - %s,
                    closing_stock = closing_stock + %s
                WHERE id=%s
            """, (old_qty, old_qty, product_id))

        new_date = request.form.get("date")
        new_type = request.form.get("transaction_type")
        new_qty = float(request.form.get("qty") or 0)
        new_rate = float(request.form.get("rate") or 0)
        new_amount = round(new_qty * new_rate, 2)
        remarks = request.form.get("remarks", "")

        # apply new stock effect
        if new_type == "Purchase":
            cur.execute("""
                UPDATE lube_stock
                SET purchase_qty = purchase_qty + %s,
                    closing_stock = closing_stock + %s
                WHERE id=%s
            """, (new_qty, new_qty, product_id))

        elif new_type == "Sale":
            cur.execute("""
                UPDATE lube_stock
                SET sale_qty = sale_qty + %s,
                    closing_stock = closing_stock - %s
                WHERE id=%s
            """, (new_qty, new_qty, product_id))

        cur.execute("""
            UPDATE lube_transactions
            SET date=%s,
                transaction_type=%s,
                qty=%s,
                rate=%s,
                amount=%s,
                remarks=%s
            WHERE id=%s
        """, (
            new_date,
            new_type,
            new_qty,
            new_rate,
            new_amount,
            remarks,
            id
        ))

        log_activity(
            cur, "Lube Stock", "Updated",
            f"Edited transaction #{id} ({old['product_name']}) — qty {old_qty}→{new_qty}, type {old_type}→{new_type}"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("lube_stock"))

@app.route("/update-lube/<int:id>", methods=["POST"])
def update_lube(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE lube_stock
            SET
                product_name=%s,
                selling_rate=%s,
                opening_stock=%s,
                purchase_qty=%s,
                sale_qty=%s,
                closing_stock=%s
            WHERE id=%s
        """, (

            request.form.get("product_name"),
            request.form.get("selling_rate"),
            request.form.get("opening_stock"),
            request.form.get("purchase_qty"),
            request.form.get("sale_qty"),
            request.form.get("closing_stock"),
            id
        ))

        # keep past transactions showing the current product name
        # after a rename, instead of freezing on the old one
        cur.execute("""
            UPDATE lube_transactions
            SET product_name=%s
            WHERE product_id=%s
        """, (
            request.form.get("product_name"),
            id
        ))

        log_activity(
            cur, "Lube Stock", "Updated",
            f"Edited product #{id} ('{request.form.get('product_name')}')"
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return redirect(url_for("lube_stock"))
# =========================================
# ATTENDANCE DATABASE UPGRADE
# put this inside init_db(), after staff_master table
# =========================================




# =========================================
# ATTENDANCE PAGE
# =========================================

@app.route("/attendance")
def attendance():

    if not session.get("logged_in"):
        return redirect(url_for("login"))
    
    selected_date = request.args.get(
    "date",
    datetime.now().strftime("%Y-%m-%d")
)

    conn = get_pg_conn()
    cur = conn.cursor()

    current_date = datetime.now().strftime("%Y-%m-%d")

    cur.execute("""
        SELECT *
        FROM staff_master
        ORDER BY id DESC
    """)
    staff_list = cur.fetchall()

    cur.execute("""
        SELECT *
        FROM attendance
        ORDER BY date DESC, id DESC
    """)
    attendance_list = cur.fetchall()

    cur.execute("""
        SELECT *
        FROM salary_payments
        ORDER BY payment_date DESC, id DESC
    """)
    salaries = cur.fetchall()

    total_staff = len(staff_list)

    present_today = 0
    absent_today = 0
    leave_today = 0

    for row in attendance_list:
        if str(row["date"]) == current_date:
            if row["attendance_status"] == "Present":
                present_today += 1
            elif row["attendance_status"] == "Absent":
                absent_today += 1
            elif row["attendance_status"] == "Leave":
                leave_today += 1

    staff_attendance = {}

    for staff in staff_list:

        staff_name = staff["staff_name"]

        staff_attendance[staff_name] = {
            "present": 0,
            "absent": 0,
            "leave": 0,
            "records": []
        }

        for att in attendance_list:

            if att["staff_name"] == staff_name:

                staff_attendance[staff_name]["records"].append(att)

                if att["attendance_status"] == "Present":
                    staff_attendance[staff_name]["present"] += 1

                elif att["attendance_status"] == "Absent":
                    staff_attendance[staff_name]["absent"] += 1

                elif att["attendance_status"] == "Leave":
                    staff_attendance[staff_name]["leave"] += 1

    salary_by_staff = {}

    for s in salaries:

        employee_name = s["employee_name"]

        if employee_name not in salary_by_staff:
            salary_by_staff[employee_name] = []

        salary_by_staff[employee_name].append(s)

    cur.execute("""
        SELECT emp_id
        FROM staff_master
        ORDER BY id DESC
        LIMIT 1
    """)
    last_emp = cur.fetchone()

    if last_emp and last_emp["emp_id"]:
        try:
            last_num = int(last_emp["emp_id"].replace("EMP", ""))
            next_emp_id = f"EMP{last_num + 1:03d}"
        except:
            next_emp_id = "EMP001"
    else:
        next_emp_id = "EMP001"

    conn.close()

    return render_template(
        "attendance.html",
        staff_list=staff_list,
        attendance_list=attendance_list,
        staff_attendance=staff_attendance,
        current_date=current_date,
        total_staff=total_staff,
        present_today=present_today,
        absent_today=absent_today,
        leave_today=leave_today,
        next_emp_id=next_emp_id,
        salaries=salaries,
        selected_date=selected_date,
        salary_by_staff=salary_by_staff
    )


@app.route("/update-salary/<int:id>", methods=["POST"])
def update_salary(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        UPDATE salary_payments
        SET
            payment_date=%s,
            month_name=%s,
            employee_name=%s,
            payment_mode=%s,
            bank_account=%s,
            amount=%s,
            remarks=%s
        WHERE id=%s
    """, (

        request.form["payment_date"],
        request.form["month_name"],
        request.form["employee_name"],
        request.form["payment_mode"],
        request.form.get("bank_account",""),
        float(request.form["amount"]),
        request.form.get("remarks",""),
        id

    ))

    log_activity(
        cur, "Attendance", "Updated",
        f"Edited salary payment #{id} for {request.form['employee_name']} — ₹{request.form['amount']}"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("attendance"))

@app.route("/edit-salary/<int:id>")
def edit_salary(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM salary_payments WHERE id=%s",
        (id,)
    )

    salary = cur.fetchone()

    conn.close()

    return render_template(
        "edit_salary.html",
        salary=salary
    )

@app.route("/delete-salary/<int:id>")
def delete_salary(id):

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT employee_name, amount FROM salary_payments WHERE id=%s", (id,))
    sal = cur.fetchone()

    cur.execute(
        "DELETE FROM salary_payments WHERE id=%s",
        (id,)
    )

    if sal:
        log_activity(
            cur, "Attendance", "Deleted",
            f"Deleted salary payment of ₹{sal['amount']} for {sal['employee_name']}"
        )

    conn.commit()
    conn.close()

    return redirect(url_for("attendance"))

@app.route("/export-attendance-excel")
def export_attendance_excel():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    wb = Workbook()

    # ==========================
    # SHEET 1 : ATTENDANCE SUMMARY
    # ==========================

    ws1 = wb.active
    ws1.title = "Attendance Summary"

    ws1.append([
        "Employee",
        "Present",
        "Absent",
        "Leave",
        "Salary Paid"
    ])

    for cell in ws1[1]:
        cell.font = Font(bold=True)

    cur.execute("""
        SELECT *
        FROM staff_master
        ORDER BY staff_name
    """)
    employees = cur.fetchall()

    for emp in employees:

        staff_name = emp["staff_name"]

        cur.execute("""
            SELECT attendance_status
            FROM attendance
            WHERE staff_name=%s
        """, (staff_name,))

        records = cur.fetchall()

        present = 0
        absent = 0
        leave = 0

        for r in records:

            if r["attendance_status"] == "Present":
                present += 1

            elif r["attendance_status"] == "Absent":
                absent += 1

            elif r["attendance_status"] == "Leave":
                leave += 1

        cur.execute("""
            SELECT SUM(amount) total_salary
            FROM salary_payments
            WHERE employee_name=%s
        """, (staff_name,))

        sal = cur.fetchone()

        salary_paid = (
            sal["total_salary"]
            if sal and sal["total_salary"]
            else 0
        )

        ws1.append([
            staff_name,
            present,
            absent,
            leave,
            salary_paid
        ])

    # ==========================
    # SHEET 2 : ATTENDANCE DETAILS
    # ==========================

    ws2 = wb.create_sheet("Attendance Details")

    ws2.append([
        "Date",
        "Employee",
        "Status"
    ])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    cur.execute("""
        SELECT *
        FROM attendance
        ORDER BY date DESC
    """)

    attendance_rows = cur.fetchall()

    for row in attendance_rows:

        ws2.append([
            row["date"],
            row["staff_name"],
            row["attendance_status"]
        ])

    # ==========================
    # SHEET 3 : SALARY HISTORY
    # ==========================

    ws3 = wb.create_sheet("Salary Payments")

    ws3.append([
        "Date",
        "Month",
        "Employee",
        "Mode",
        "Bank Account",
        "Amount",
        "Remarks"
    ])

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    cur.execute("""
        SELECT *
        FROM salary_payments
        ORDER BY payment_date DESC
    """)

    salaries = cur.fetchall()

    for s in salaries:

        ws3.append([
            s["payment_date"],
            s["month_name"],
            s["employee_name"],
            s["payment_mode"],
            s["bank_account"],
            s["amount"],
            s["remarks"]
        ])

    # ==========================
    # EMPLOYEE-WISE SHEETS
    # ==========================

    for emp in employees:

        staff_name = emp["staff_name"]

        sheet_name = staff_name[:31]

        ws = wb.create_sheet(sheet_name)

        ws.append([
            "Date",
            "Attendance Status"
        ])

        for cell in ws[1]:
            cell.font = Font(bold=True)

        cur.execute("""
            SELECT *
            FROM attendance
            WHERE staff_name=%s
            ORDER BY date
        """, (staff_name,))

        emp_attendance = cur.fetchall()

        present = 0
        absent = 0
        leave = 0

        for a in emp_attendance:

            ws.append([
                a["date"],
                a["attendance_status"]
            ])

            if a["attendance_status"] == "Present":
                present += 1

            elif a["attendance_status"] == "Absent":
                absent += 1

            elif a["attendance_status"] == "Leave":
                leave += 1

        cur.execute("""
            SELECT SUM(amount) total_salary
            FROM salary_payments
            WHERE employee_name=%s
        """, (staff_name,))

        sal = cur.fetchone()

        salary_paid = (
            sal["total_salary"]
            if sal and sal["total_salary"]
            else 0
        )

        ws.append([])
        ws.append(["Present", present])
        ws.append(["Absent", absent])
        ws.append(["Leave", leave])
        ws.append(["Salary Paid", salary_paid])

    conn.close()

    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
    )

    wb.save(temp_file.name)

    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name="Attendance_Report.xlsx"
    )

# =========================================
# SAVE STAFF
# =========================================

@app.route("/save-staff", methods=["POST"])
def save_staff():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn, cur = get_pg_cursor()

    staff_name = request.form.get(
        "staff_name",
        ""
    ).strip()

    if not staff_name:

        conn.close()

        return redirect(
            url_for("attendance")
        )

    cur.execute("""
        SELECT id
        FROM staff_master
        WHERE LOWER(staff_name)=LOWER(%s)
    """, (staff_name,))

    existing = cur.fetchone()

    if not existing:

        cur.execute("""

            INSERT INTO staff_master (

                emp_id,
                staff_name,
                role,
                department,
                joined_date,
                bank_account,
                shift,
                status

            )

            VALUES (

                %s,%s,%s,%s,%s,%s,%s,%s

            )

        """, (

            request.form.get("emp_id", ""),

            staff_name,

            request.form.get("role", ""),

            request.form.get("department", ""),

            request.form.get("joined_date", None),

            request.form.get("bank_account", ""),

            request.form.get("shift", ""),

            request.form.get("status", "Active")

        ))

        log_activity(
            cur, "Attendance", "Created",
            f"Added staff '{staff_name}' ({request.form.get('role','')})"
        )

    conn.commit()

    conn.close()

    return redirect(
        url_for("attendance")
    )

# =========================================
# SAVE / UPDATE ATTENDANCE
# =========================================




@app.route("/save-attendance", methods=["POST"])
def save_attendance():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    attendance_date = request.form.get(
        "date",
        datetime.now().strftime("%Y-%m-%d")
    )

    staff_name = request.form.get(
        "staff_name",
        ""
    ).strip()

    attendance_status = request.form.get(
        "attendance_status",
        ""
    )

    if not staff_name or not attendance_status:
        return redirect(
            url_for("attendance", date=attendance_date)
        )

    selected = datetime.strptime(
        attendance_date,
        "%Y-%m-%d"
    )

    if selected.date() > datetime.now().date():

        flash("Future attendance not allowed", "error")

        return redirect(
            url_for("attendance", date=attendance_date)
        )

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id
        FROM attendance
        WHERE date=%s AND staff_name=%s
        LIMIT 1
    """, (
        attendance_date,
        staff_name
    ))

    existing = cur.fetchone()

    if existing:

        cur.execute("""
            UPDATE attendance
            SET attendance_status=%s
            WHERE id=%s
        """, (
            attendance_status,
            existing["id"]
        ))

    else:

        cur.execute("""
            INSERT INTO attendance (
                date,
                staff_name,
                attendance_status
            )
            VALUES (%s, %s, %s)
        """, (
            attendance_date,
            staff_name,
            attendance_status
        ))

    log_activity(
        cur, "Attendance",
        "Updated" if existing else "Created",
        f"Marked {staff_name} as {attendance_status} on {attendance_date}"
    )

    conn.commit()
    conn.close()

    return redirect(
        url_for("attendance", date=attendance_date)
    )

@app.route("/delete-lube/<int:id>")
def delete_lube(id):
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("SELECT product_name FROM lube_stock WHERE id=%s", (id,))
    prod = cur.fetchone()

    cur.execute("DELETE FROM lube_stock WHERE id = %s", (id,))

    if prod:
        log_activity(
            cur, "Lube Stock", "Deleted",
            f"Deleted product '{prod['product_name']}'"
        )

    conn.commit()
    conn.close()

    return redirect(url_for("lube_stock"))

@app.route("/full-system-export")
def full_system_export():

    if not session.get("logged_in") or session.get("role") != "admin":
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="07120C")
    sub_fill = PatternFill("solid", fgColor="16A34A")
    light_fill = PatternFill("solid", fgColor="F1F5F9")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=18, bold=True, color="07120C")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    def style_sheet(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                cell.border = border

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except:
                    pass

            ws.column_dimensions[col_letter].width = min(max_len + 4, 32)

    def export_table(sheet_name, table_name):
        cur.execute(f"SELECT * FROM {table_name}")
        rows = cur.fetchall()

        ws = wb.create_sheet(sheet_name)

        if not rows:
            ws.append(["No Data"])
            return ws

        headers = rows[0].keys()
        ws.append(list(headers))

        for r in rows:
            ws.append([r[h] for h in headers])

        style_sheet(ws)
        ws.freeze_panes = "A2"

        return ws

    # =========================
    # DASHBOARD SHEET
    # =========================

    ws = wb.active
    ws.title = "Dashboard"

    ws.merge_cells("A1:H1")
    ws["A1"] = "SAI FUEL MART - FULL SYSTEM EXPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = light_fill

    cur.execute("""
        SELECT
            COALESCE(SUM(total_fuel_sale),0) AS fuel_sale,
            COALESCE(SUM(lube_sale),0) AS lube_sale,
            COALESCE(SUM(digital_collection),0) AS digital,
            COALESCE(SUM(credit_given),0) AS credit,
            COALESCE(SUM(total_expense),0) AS expense,
            COALESCE(SUM(cash_in_hand),0) AS cash
        FROM daily_closing
    """)
    d = cur.fetchone()

    cur.execute("SELECT COUNT(*) AS total FROM credit_transporters")
    total_transporters = cur.fetchone()["total"]

    cur.execute("SELECT COUNT(*) AS total FROM staff_master")
    total_staff = cur.fetchone()["total"]

    cur.execute("SELECT COALESCE(SUM(balance_due),0) AS total FROM credit_transporters")
    total_due = cur.fetchone()["total"]

    dashboard_data = [
        ["Metric", "Value"],
        ["Total Fuel Sale", d["fuel_sale"]],
        ["Total Lube Sale", d["lube_sale"]],
        ["Digital Collection", d["digital"]],
        ["Total Credit Given", d["credit"]],
        ["Total Expense", d["expense"]],
        ["Cash In Hand", d["cash"]],
        ["Total Transporters", total_transporters],
        ["Total Staff", total_staff],
        ["Outstanding Credit", total_due],
        ["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]

    start_row = 3
    for r in dashboard_data:
        ws.append(r)

    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = white_font

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # =========================
    # MONTHLY TREND
    # =========================

    cur.execute("""
    SELECT
        TO_CHAR(date, 'YYYY-MM') AS month,
        ROUND(SUM(total_fuel_sale),2) AS fuel,
        ROUND(SUM(lube_sale),2) AS lube,
        ROUND(SUM(credit_given),2) AS credit
    FROM daily_closing
    GROUP BY TO_CHAR(date, 'YYYY-MM')
    ORDER BY TO_CHAR(date, 'YYYY-MM')
""")
    monthly = cur.fetchall()

    ws["D3"] = "Monthly Trend"
    ws["D3"].font = Font(bold=True, size=14)

    ws.append([])
    ws["D5"] = "Month"
    ws["E5"] = "Fuel Sale"
    ws["F5"] = "Lube Sale"
    ws["G5"] = "Credit"

    for c in ["D5", "E5", "F5", "G5"]:
        ws[c].fill = sub_fill
        ws[c].font = white_font

    row_no = 6
    for m in monthly:
        ws[f"D{row_no}"] = m["month"]
        ws[f"E{row_no}"] = m["fuel"]
        ws[f"F{row_no}"] = m["lube"]
        ws[f"G{row_no}"] = m["credit"]
        row_no += 1

    if row_no > 6:
        chart = LineChart()
        chart.title = "Monthly Fuel / Lube / Credit Trend"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Month"

        data_ref = Reference(ws, min_col=5, max_col=7, min_row=5, max_row=row_no-1)
        cats = Reference(ws, min_col=4, min_row=6, max_row=row_no-1)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18

        ws.add_chart(chart, "D14")

    # =========================
    # SHEETS FROM TABLES
    # =========================

    tables = [
        ("Daily Closing", "daily_closing"),
        ("Nozzle Entries", "nozzle_entries"),
        ("Nozzle Master", "nozzle_master"),
        ("Tank Level", "tank_level"),
        ("Credit Transporters", "credit_transporters"),
        ("Transport Entries", "transport_entries"),
        ("Transport Ledger", "transporter_ledger"),
        ("Attendance", "attendance"),
        ("Staff Master", "staff_master"),
        ("Salary Payments", "salary_payments"),
        ("Lube Stock", "lube_stock"),
        ("Lube Transactions", "lube_transactions"),
        ("Settings", "settings")
    ]

    for sheet_name, table_name in tables:
        try:
            export_table(sheet_name, table_name)
        except Exception:
            ws_err = wb.create_sheet(sheet_name)
            ws_err.append(["Table not found or no access", table_name])

    # =========================
    # ATTENDANCE SUMMARY SHEET
    # =========================

    ws_att = wb.create_sheet("Attendance Summary")
    ws_att.append(["Employee", "Present", "Absent", "Leave", "Salary Paid"])

    cur.execute("SELECT * FROM staff_master ORDER BY staff_name")
    staff_rows = cur.fetchall()

    for staff in staff_rows:
        name = staff["staff_name"]

        cur.execute("""
            SELECT attendance_status
            FROM attendance
            WHERE staff_name=%s
        """, (name,))
        records = cur.fetchall()

        present = sum(1 for r in records if r["attendance_status"] == "Present")
        absent = sum(1 for r in records if r["attendance_status"] == "Absent")
        leave = sum(1 for r in records if r["attendance_status"] == "Leave")

        cur.execute("""
            SELECT COALESCE(SUM(amount),0) AS total
            FROM salary_payments
            WHERE employee_name=%s
        """, (name,))
        salary = cur.fetchone()["total"]

        ws_att.append([name, present, absent, leave, salary])

    style_sheet(ws_att)

    # =========================
    # TRANSPORT SUMMARY SHEET
    # =========================

    ws_tr = wb.create_sheet("Transport Summary")
    ws_tr.append(["Party", "Fuel Credit", "Lube Credit", "Total Credit", "Received", "Balance"])

    cur.execute("""
        SELECT party_name, fuel_credit, lube_credit,
               credit_given, payment_received, balance_due
        FROM credit_transporters
        ORDER BY balance_due DESC
    """)
    trs = cur.fetchall()

    for t in trs:
        ws_tr.append([
            t["party_name"],
            t["fuel_credit"],
            t["lube_credit"],
            t["credit_given"],
            t["payment_received"],
            t["balance_due"]
        ])

    style_sheet(ws_tr)

    if len(trs) > 0:
        pie = PieChart()
        pie.title = "Outstanding Credit by Transporter"

        labels = Reference(ws_tr, min_col=1, min_row=2, max_row=len(trs)+1)
        data = Reference(ws_tr, min_col=6, min_row=1, max_row=len(trs)+1)

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 8
        pie.width = 12

        ws_tr.add_chart(pie, "H2")

    # =========================
    # LUBE SUMMARY SHEET
    # =========================

    ws_lube = wb.create_sheet("Lube Summary")
    ws_lube.append(["Product", "Opening", "Purchase", "Sale", "Closing", "Rate", "Closing Value"])

    cur.execute("""
        SELECT *
        FROM lube_stock
        ORDER BY product_name
    """)
    lube_rows = cur.fetchall()

    for l in lube_rows:
        closing_value = float(l["closing_stock"] or 0) * float(l["selling_rate"] or 0)

        ws_lube.append([
            l["product_name"],
            l["opening_stock"],
            l["purchase_qty"],
            l["sale_qty"],
            l["closing_stock"],
            l["selling_rate"],
            closing_value
        ])

    style_sheet(ws_lube)

    # =========================
    # CHART DASHBOARD EXTRA
    # =========================

    ws_chart = wb.create_sheet("Charts")

    ws_chart.append(["Chart Type", "Description"])
    ws_chart.append(["Line Chart", "Monthly fuel, lube and credit trend is on Dashboard sheet"])
    ws_chart.append(["Pie Chart", "Outstanding credit by transporter is on Transport Summary sheet"])
    style_sheet(ws_chart)

    # =========================
    # FORMAT ALL SHEETS
    # =========================

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")

    conn.close()

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)

    file_name = "Sai_Fuel_Mart_Full_System_Export.xlsx"

    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=file_name
    )

@app.route("/analytics")
def analytics():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    conn = get_pg_conn()
    cur = conn.cursor()

    current_month = request.args.get(
    "month",
    datetime.now().strftime("%Y-%m")
)

    # =========================
    # TODAY SALES
    # =========================

    cur.execute("""
SELECT
COALESCE(SUM(total_fuel_sale),0) fuel_sale,
COALESCE(SUM(lube_sale),0) lube_sale,
COALESCE(SUM(cash_in_hand),0) cash_in_hand,
COALESCE(SUM(total_expense),0) expense
FROM daily_closing
WHERE TO_CHAR(date, 'YYYY-MM')=%s
""",(current_month,))

    sales = cur.fetchone()

    fuel_sale = float(sales["fuel_sale"] or 0)
    lube_sale = float(sales["lube_sale"] or 0)
    cash_in_hand = float(sales["cash_in_hand"] or 0)
    total_expense = float(sales["expense"] or 0)

    today_sale = fuel_sale + lube_sale

    net_profit = today_sale - total_expense

    # =========================
    # CREDIT TRANSPORT
    # =========================

    cur.execute("""
        SELECT
            COUNT(*) total_transporters,
            COALESCE(SUM(credit_given),0) total_credit,
            COALESCE(SUM(payment_received),0) total_received,
            COALESCE(SUM(balance_due),0) outstanding
        FROM credit_transporters
    """)

    transport = cur.fetchone()

    # =========================
    # STAFF
    # =========================

    cur.execute("""
        SELECT COUNT(*) total_staff
        FROM staff_master
        WHERE status='Active'
    """)

    total_staff = cur.fetchone()["total_staff"]

    cur.execute("""
        SELECT attendance_status
        FROM attendance
        WHERE TO_CHAR(date, 'YYYY-MM')=%s
    """, (current_month,))

    attendance_rows = cur.fetchall()

    present_today = 0
    absent_today = 0
    leave_today = 0

    for row in attendance_rows:

        if row["attendance_status"] == "Present":
            present_today += 1

        elif row["attendance_status"] == "Absent":
            absent_today += 1

        elif row["attendance_status"] == "Leave":
            leave_today += 1

    # =========================
    # SALARY
    # =========================

    cur.execute("""
        SELECT
        COALESCE(SUM(amount),0) salary_paid
        FROM salary_payments
        WHERE TO_CHAR(payment_date, 'YYYY-MM')=%s
    """, (current_month,))

    salary_paid = cur.fetchone()["salary_paid"]

    # =========================
    # LUBE ANALYTICS
    # =========================

    cur.execute("""
        SELECT
            product_name,
            closing_stock,
            selling_rate
        FROM lube_stock
    """)

    lube_rows = cur.fetchall()

    total_lube_value = 0

    for item in lube_rows:

        stock = float(item["closing_stock"] or 0)
        rate = float(item["selling_rate"] or 0)

        total_lube_value += stock * rate

    cur.execute("""
        SELECT
        product_name,
        sale_qty
        FROM lube_stock
        ORDER BY sale_qty DESC
        LIMIT 1
    """)

    top_product = cur.fetchone()

    top_product_name = (
        top_product["product_name"]
        if top_product
        else "-"
    )

    # =========================
    # LOW STOCK ALERTS
    # =========================

    cur.execute("""
        SELECT COUNT(*) low_stock
        FROM lube_stock
        WHERE closing_stock < 10
    """)

    low_stock = cur.fetchone()["low_stock"]

    # =========================
    # TOP CREDITORS
    # =========================

    cur.execute("""
        SELECT
            party_name,
            balance_due
        FROM credit_transporters
        ORDER BY balance_due DESC
        LIMIT 10
    """)

    top_creditors = cur.fetchall()

    # =========================
    # MONTHLY TREND
    # =========================

    cur.execute("""
        SELECT
            TO_CHAR(date, 'YYYY-MM') AS month,
            ROUND(SUM(total_fuel_sale),2) fuel_sale,
            ROUND(SUM(lube_sale),2) lube_sale
        FROM daily_closing
        GROUP BY TO_CHAR(date, 'YYYY-MM')
        ORDER BY month
    """)

    monthly_trend = cur.fetchall()

    conn.close()

    import json

    trend_labels = json.dumps(
    [x["month"] for x in monthly_trend]
)

    fuel_data = json.dumps(
    [float(x["fuel_sale"] or 0) for x in monthly_trend]
)

    lube_data = json.dumps(
    [float(x["lube_sale"] or 0) for x in monthly_trend]
)

    if total_staff > 0:
     staff_percent = int((present_today * 100) / total_staff)
    else:
     staff_percent = 0


    return render_template(
        "analytics.html",

        today_sale=today_sale,
        fuel_sale=fuel_sale,
        lube_sale=lube_sale,

        cash_in_hand=cash_in_hand,
        total_expense=total_expense,
        net_profit=net_profit,

        total_transporters=transport["total_transporters"],
        total_credit=transport["total_credit"],
        total_received=transport["total_received"],
        outstanding_credit=transport["outstanding"],

        total_staff=total_staff,
        present_today=present_today,
        absent_today=absent_today,
        leave_today=leave_today,

        salary_paid=salary_paid,

        total_lube_value=total_lube_value,
        top_product_name=top_product_name,

        low_stock=low_stock,

        top_creditors=top_creditors,
        monthly_trend=monthly_trend,
        trend_labels=trend_labels,
        fuel_data=fuel_data,
        lube_data=lube_data,
        staff_percent=staff_percent,
        
    )

def get_supabase_client():

    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY")

    return create_client(url, key)

def upload_proof_file(file, folder):
    if not file or file.filename == "":
        return ""

    supabase = get_supabase_client()

    ext = file.filename.rsplit(".", 1)[-1].lower()
    file_name = f"{folder}/{uuid.uuid4()}.{ext}"

    file_bytes = file.read()

    supabase.storage.from_("proof-files").upload(
        file_name,
        file_bytes,
        {"content-type": file.content_type}
    )

    return supabase.storage.from_("proof-files").get_public_url(file_name)


@app.route("/proof-pdf")
def proof_pdf():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    proof_date = request.args.get("date")
    category = request.args.get("category")

    if not proof_date or not category:
        flash("Date and category are required to generate a proof report.")
        return redirect(url_for("proof_register"))

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM proof_register
        WHERE proof_date=%s AND proof_category=%s
        ORDER BY fuel_type ASC, item_name ASC
    """, (proof_date, category))
    rows = cur.fetchall()

    cur.execute("SELECT * FROM settings WHERE id=1")
    biz = cur.fetchone()

    conn.close()

    station_name = (biz["station_name"] if biz and biz["station_name"] else "") or "SAI FUEL MART"
    station_address = biz["station_address"] if biz and biz["station_address"] else ""

    file = BytesIO()

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#e5e7eb"))
        canvas.line(30, 30, 565, 30)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawRightString(565, 18, f"Page {doc.page}")
        canvas.drawString(30, 18, f"{station_name} — {category} Proof Report — {proof_date}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        file,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ProofTitle", parent=styles["Title"],
        fontSize=18, leading=20, textColor=colors.HexColor("#07120C"), alignment=0
    )
    sub_style = ParagraphStyle(
        "ProofSub", parent=styles["Normal"],
        fontSize=9.5, leading=13, textColor=colors.HexColor("#475467")
    )
    band_style = ParagraphStyle(
        "Band", parent=styles["Normal"],
        fontSize=12, leading=14, textColor=colors.white, fontName="Helvetica-Bold"
    )
    meta_style = ParagraphStyle(
        "Meta", parent=styles["Normal"],
        fontSize=9.5, leading=14, textColor=colors.HexColor("#07120C")
    )
    link_style = ParagraphStyle(
        "Link", parent=meta_style,
        textColor=colors.HexColor("#2563eb")
    )
    empty_style = ParagraphStyle(
        "Empty", parent=styles["Normal"],
        fontSize=11, textColor=colors.HexColor("#94a3b8"), alignment=1
    )

    elements = []

    elements.append(Paragraph(station_name.upper(), title_style))
    if station_address:
        elements.append(Paragraph(station_address, sub_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"<b>{category} — Proof Report</b>", sub_style))
    elements.append(Paragraph(f"Date: {proof_date} &nbsp;|&nbsp; Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", sub_style))
    elements.append(Spacer(1, 6))

    accent = Table([[""]], colWidths=[535], rowHeights=[3])
    accent.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#16a34a"))]))
    elements.append(accent)
    elements.append(Spacer(1, 14))

    if not rows:
        elements.append(Spacer(1, 40))
        elements.append(Paragraph("No proof entries were uploaded for this date and category.", empty_style))

    current_fuel = None

    for r in rows:

        if r["fuel_type"] != current_fuel:
            current_fuel = r["fuel_type"]
            band = Table([[Paragraph(current_fuel or "General", band_style)]], colWidths=[535])
            band.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#07120C")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ]))
            elements.append(Spacer(1, 6))
            elements.append(band)
            elements.append(Spacer(1, 8))

        meta_line = f"<b>{r['item_name']}</b> &nbsp;|&nbsp; Status: {r['stock_status']} &nbsp;|&nbsp; Time: {r['proof_time']}"
        if r["latitude"] and r["longitude"]:
            try:
                meta_line += f" &nbsp;|&nbsp; GPS: {float(r['latitude']):.5f}, {float(r['longitude']):.5f}"
            except (TypeError, ValueError):
                pass

        elements.append(Paragraph(meta_line, meta_style))
        elements.append(Spacer(1, 4))

        if r["stock_status"] == "No Stock":
            elements.append(Paragraph(f"<i>{r['remarks'] or 'No stock reported'}</i>", meta_style))

        elif r["photo_url"]:
            try:
                resp = requests.get(r["photo_url"], timeout=15)
                img_stream = BytesIO(resp.content)
                img = Image(img_stream, width=210, height=157)
                elements.append(img)
            except Exception:
                elements.append(Paragraph("(photo could not be loaded for this report)", meta_style))

        elif r["video_url"]:
            elements.append(Paragraph(
                f'<link href="{r["video_url"]}">&#9654; Watch recorded video</link>',
                link_style
            ))

        elements.append(Spacer(1, 16))

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)

    file.seek(0)

    safe_date = str(proof_date).replace("-", "")
    safe_category = category.replace(" ", "_")
    filename = f"{safe_category}_{safe_date}.pdf"

    as_attachment = request.args.get("download") == "1"

    return send_file(
        file,
        as_attachment=as_attachment,
        download_name=filename,
        mimetype="application/pdf"
    )


@app.route("/proof-upload")
def proof_upload():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    return render_template(
        "proof_upload.html",
        today=datetime.now().strftime("%Y-%m-%d")
    )




@app.route("/proof-register")
def proof_register():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    category = request.args.get("category", "")
    fuel_type = request.args.get("fuel_type", "")

    conn = get_pg_conn()
    cur = conn.cursor()

    query = """
        SELECT *
        FROM proof_register
        WHERE 1=1
    """

    params = []

    if from_date:
        query += " AND proof_date >= %s"
        params.append(from_date)

    if to_date:
        query += " AND proof_date <= %s"
        params.append(to_date)

    if category:
        query += " AND proof_category = %s"
        params.append(category)

    if fuel_type:
        query += " AND fuel_type = %s"
        params.append(fuel_type)

    query += """
        ORDER BY proof_date DESC, proof_time DESC, id DESC
    """

    cur.execute(query, params)
    rows = cur.fetchall()

    # distinct (date, category) groups present in this filtered view,
    # so the register can offer one "View/Download PDF" per group
    seen = set()
    groups = []
    for r in rows:
        key = (r["proof_date"], r["proof_category"])
        if key not in seen:
            seen.add(key)
            groups.append({"date": r["proof_date"], "category": r["proof_category"]})

    conn.close()

    return render_template(
        "proof_register.html",
        rows=rows,
        groups=groups
    )


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


## init_db()

@app.route("/download-db")
def download_db():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    if not is_admin():
        return redirect(url_for("dashboard"))

    # same fix as /backup-database — the real data lives in Postgres,
    # never in this local file
    return redirect(url_for("full_system_export"))

def get_pg_conn():

    return psycopg2.connect(
        os.environ["DATABASE_URL"],
        cursor_factory=psycopg2.extras.RealDictCursor,
        sslmode="require"
    )

def get_pg_cursor():

    conn = get_pg_conn()

    cur = conn.cursor()

    return conn, cur


def ensure_activity_log_table():
    """
    Make sure the audit-log table exists. Runs once at startup — safe to
    call repeatedly since it's IF NOT EXISTS, and won't crash the app if
    the database isn't reachable yet at import time.
    """
    try:
        conn = get_pg_conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS activity_log (
                id SERIAL PRIMARY KEY,
                logged_at TIMESTAMP DEFAULT NOW(),
                username TEXT,
                role TEXT,
                module TEXT,
                action TEXT,
                description TEXT
            )
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[activity_log] could not ensure table exists: {e}")


ensure_activity_log_table()


def ensure_settings_columns():
    """
    Make sure the settings table has the Business Info columns used by
    the redesigned Settings page and the transporter bill PDF letterhead.
    Runs once at startup, safe to call repeatedly, won't crash the app
    if the database isn't reachable yet at import time.
    """
    try:
        conn = get_pg_conn()
        cur = conn.cursor()
        for col in ["station_name", "station_address", "gstin", "phone_number"]:
            cur.execute(f"""
                ALTER TABLE settings ADD COLUMN IF NOT EXISTS {col} TEXT
            """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[settings] could not ensure business-info columns exist: {e}")


ensure_settings_columns()


def ensure_transporter_discount_column():
    """
    Make sure credit_transporters has a per-transporter discount column
    (flat Rs. amount, not a percentage), used by Settings and applied at
    the bottom of the transporter bill PDF/Excel export.
    """
    try:
        conn = get_pg_conn()
        cur = conn.cursor()
        cur.execute("""
            ALTER TABLE credit_transporters ADD COLUMN IF NOT EXISTS discount REAL DEFAULT 0
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[credit_transporters] could not ensure discount column exists: {e}")


ensure_transporter_discount_column()


def ensure_tank_level_atg_columns():
    """
    Make sure tank_level has every field the HP Smart Connect ATG page
    shows — used for the animated tank visual, alerts, and the detailed
    live-readings panel on Tank Level.
    """
    try:
        conn = get_pg_conn()
        cur = conn.cursor()
        for col, coltype in [
            ("capacity", "REAL DEFAULT 0"),
            ("water_ltr", "REAL DEFAULT 0"),
            ("temperature_c", "REAL DEFAULT 0"),
            ("dip_mm", "REAL DEFAULT 0"),
            ("gross_volume", "REAL DEFAULT 0"),
            ("water_dip_mm", "REAL DEFAULT 0"),
            ("ullage_ltr", "REAL DEFAULT 0"),
            ("decantation_status", "TEXT"),
            ("density_status", "TEXT"),
            ("density_at_15", "REAL DEFAULT 0"),
            ("density_kg_m3", "REAL DEFAULT 0"),
            ("den_float_height", "REAL DEFAULT 0"),
        ]:
            cur.execute(f"ALTER TABLE tank_level ADD COLUMN IF NOT EXISTS {col} {coltype}")
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[tank_level] could not ensure ATG columns exist: {e}")


ensure_tank_level_atg_columns()


def ensure_fuel_receipts_table():
    """
    Table for fuel delivery invoices/challans (like the HPCL delivery
    receipt) — a separate record from day-to-day Tank Level, which is
    now driven automatically by the ATG bookmarklet sync. One receipt
    covers the whole vehicle: shared header fields, then each of the
    5 compartments independently tagged MS or HSD with its own dip/vol.
    """
    try:
        conn = get_pg_conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fuel_receipts (
                id SERIAL PRIMARY KEY,
                receipt_date TEXT,
                invoice_no TEXT,
                order_no TEXT,
                vehicle_no TEXT,
                carrier_no TEXT,
                carrier_name TEXT,
                po_no TEXT,
                po_date TEXT,
                water_checked TEXT,
                density REAL DEFAULT 0,
                temperature_c REAL DEFAULT 0,
                dip_before REAL DEFAULT 0,
                dip_after REAL DEFAULT 0,
                comp1_fuel TEXT, comp1_dip REAL DEFAULT 0, comp1_vol REAL DEFAULT 0,
                comp2_fuel TEXT, comp2_dip REAL DEFAULT 0, comp2_vol REAL DEFAULT 0,
                comp3_fuel TEXT, comp3_dip REAL DEFAULT 0, comp3_vol REAL DEFAULT 0,
                comp4_fuel TEXT, comp4_dip REAL DEFAULT 0, comp4_vol REAL DEFAULT 0,
                comp5_fuel TEXT, comp5_dip REAL DEFAULT 0, comp5_vol REAL DEFAULT 0,
                total_ms_vol REAL DEFAULT 0,
                total_hsd_vol REAL DEFAULT 0,
                created_at TEXT
            )
        """)

        # migrate an existing (pre-redesign) table safely: add any new
        # columns that don't exist yet, drop ones we no longer use
        for col, coltype in [
            ("comp1_fuel", "TEXT"), ("comp2_fuel", "TEXT"), ("comp3_fuel", "TEXT"),
            ("comp4_fuel", "TEXT"), ("comp5_fuel", "TEXT"),
            ("total_ms_vol", "REAL DEFAULT 0"), ("total_hsd_vol", "REAL DEFAULT 0"),
        ]:
            cur.execute(f"ALTER TABLE fuel_receipts ADD COLUMN IF NOT EXISTS {col} {coltype}")

        for col in ["photo_url", "remarks", "fuel_type", "quantity_ltr", "rate_per_unit", "total_amount"]:
            cur.execute(f"ALTER TABLE fuel_receipts DROP COLUMN IF EXISTS {col}")

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[fuel_receipts] could not ensure table exists: {e}")


ensure_fuel_receipts_table()


def log_activity(cur, module, action, description):
    """
    Record who did what, when, across every part of the app. Call this
    on the SAME cursor/transaction as the write it's logging, right
    before conn.commit() — so the log entry and the change it describes
    always land together, or not at all.
    """
    try:
        cur.execute("""
            INSERT INTO activity_log (
                username, role, module, action, description
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            session.get("username", "system"),
            session.get("role", "-"),
            module,
            action,
            description
        ))
    except Exception:
        # logging should never be able to break the actual operation
        pass

@app.route("/activity-log")
def activity_log():

    if not session.get("logged_in") or session.get("role") != "admin":
        return redirect(url_for("login"))

    module_filter = request.args.get("module", "")
    user_filter = request.args.get("username", "")
    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")

    conn = get_pg_conn()
    cur = conn.cursor()

    query = "SELECT * FROM activity_log WHERE 1=1"
    params = []

    if module_filter:
        query += " AND module=%s"
        params.append(module_filter)

    if user_filter:
        query += " AND username=%s"
        params.append(user_filter)

    if from_date:
        query += " AND logged_at::date >= %s"
        params.append(from_date)

    if to_date:
        query += " AND logged_at::date <= %s"
        params.append(to_date)

    query += " ORDER BY id DESC LIMIT 500"

    cur.execute(query, tuple(params))
    logs = cur.fetchall()

    cur.execute("SELECT DISTINCT module FROM activity_log ORDER BY module ASC")
    modules = [r["module"] for r in cur.fetchall()]

    cur.execute("SELECT DISTINCT username FROM activity_log ORDER BY username ASC")
    usernames = [r["username"] for r in cur.fetchall()]

    conn.close()

    return render_template(
        "activity_log.html",
        logs=logs,
        modules=modules,
        usernames=usernames,
        module_filter=module_filter,
        user_filter=user_filter,
        from_date=from_date,
        to_date=to_date
    )

@app.route("/delete-proof-register", methods=["POST"])
def delete_proof_register():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    proof_ids = [int(x) for x in request.form.getlist("proof_ids")]

    if not proof_ids:
        return redirect(url_for("proof_register"))

    conn = get_pg_conn()
    cur = conn.cursor()

    # get photo urls first
    cur.execute("""
        SELECT photo_url
        FROM proof_register
        WHERE id = ANY(%s)
    """, (proof_ids,))

    rows = cur.fetchall()

    supabase = get_supabase_client()

    for row in rows:
        photo_url = row["photo_url"]

        if photo_url:
            try:
                file_path = photo_url.split("/proof-files/")[-1]

                supabase.storage.from_("proof-files").remove([
                    file_path
                ])

            except Exception:
                pass

    # delete database rows
    cur.execute("""
        DELETE FROM proof_register
        WHERE id = ANY(%s)
    """, (proof_ids,))

    log_activity(
        cur, "Proof Register", "Deleted",
        f"Deleted {len(proof_ids)} proof register entr{'y' if len(proof_ids)==1 else 'ies'}"
    )

    conn.commit()
    conn.close()

    return redirect(url_for("proof_register"))

@app.route("/save-proof-upload", methods=["POST"])
def save_proof_upload():

    if not session.get("logged_in"):
        return redirect(url_for("login"))

    proof_date = request.form.get("proof_date")
    proof_time = datetime.now().strftime("%H:%M:%S")
    proof_type = request.form.get("proof_type")

    latitude = request.form.get("latitude", "")
    longitude = request.form.get("longitude", "")
    client_time = request.form.get("client_time", "")

    location_url = ""
    if latitude and longitude:
        location_url = f"https://www.google.com/maps?q={latitude},{longitude}"

    conn = get_pg_conn()
    cur = conn.cursor()

    def insert_proof(category, fuel, item, status, photo_url="", remarks="", video_url=""):
        cur.execute("""
            INSERT INTO proof_register (
                proof_date, proof_time, proof_category,
                fuel_type, item_name, stock_status,
                photo_url, video_url, remarks,
                latitude, longitude, location_url, client_time
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            proof_date,
            proof_time,
            category,
            fuel,
            item,
            status,
            photo_url,
            video_url,
            remarks,
            latitude,
            longitude,
            location_url,
            client_time
        ))

    try:

        if proof_type == "Nozzle Testing":

            for fuel in ["MS", "HSD"]:

                fuel_status = request.form.get(f"{fuel.lower()}_stock_status")
                nozzles = ["MS1", "MS2", "MS3"] if fuel == "MS" else ["HSD1", "HSD2", "HSD3"]

                if fuel_status == "No Stock":

                    for nozzle in nozzles:
                        insert_proof(
                            "Nozzle Testing",
                            fuel,
                            nozzle,
                            "No Stock",
                            "",
                            f"{fuel} no stock"
                        )

                else:

                    for nozzle in nozzles:
                        photo = request.files.get(f"{nozzle}_photo")

                        if not photo or not photo.filename:
                            conn.close()
                            return f"{nozzle} photo required"

                        photo_url = upload_proof_file(photo, "photos")

                        insert_proof(
                            "Nozzle Testing",
                            fuel,
                            nozzle,
                            "Available",
                            photo_url,
                            ""
                        )

                    # one live video per fuel type, covering all 3 machines
                    video = request.files.get(f"{fuel}_video")

                    if not video or not video.filename:
                        conn.close()
                        return f"{fuel} video required"

                    video_url = upload_proof_file(video, "videos")

                    insert_proof(
                        "Nozzle Testing",
                        fuel,
                        f"{fuel} Video",
                        "Available",
                        "",
                        "",
                        video_url
                    )

        elif proof_type == "Dip Check":

            for fuel in ["MS", "HSD"]:

                dip_status = request.form.get(f"{fuel.lower()}_dip_status")

                if dip_status == "No Stock":

                    insert_proof(
                        "Dip Check",
                        fuel,
                        f"{fuel} Dip",
                        "No Stock",
                        "",
                        f"{fuel} no stock"
                    )

                else:

                    photo = request.files.get(f"{fuel}_dip_photo")

                    if not photo or not photo.filename:
                        conn.close()
                        return f"{fuel} dip photo required"

                    photo_url = upload_proof_file(photo, "photos")

                    insert_proof(
                        "Dip Check",
                        fuel,
                        f"{fuel} Dip",
                        "Available",
                        photo_url,
                        ""
                    )

        else:
            conn.close()
            return "Please select proof type"

        log_activity(
            cur, "Proof Register", "Created",
            f"Uploaded {proof_type} proof for {proof_date}"
        )

        conn.commit()
        conn.close()

        return redirect(url_for("proof_register"))

    except Exception as e:
        conn.rollback()
        conn.close()
        return f"Proof upload error: {str(e)}"

@app.route("/api/transporters")
def api_transporters():

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM credit_transporters
        ORDER BY party_name ASC
    """)

    rows = cur.fetchall()

    conn.close()

    return jsonify(rows)

@app.route("/api/transport-reports")
def transport_reports():

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            entry_date,
            transporter_name,
            vehicle_no,
            qty,
            rate,
            total_amount
        FROM transport_entries
        ORDER BY entry_date DESC
    """)

    rows = cur.fetchall()

    data = []

    for r in rows:

        data.append({
            "entry_date": str(r["entry_date"]),
            "transporter_name": r["transporter_name"],
            "vehicle_no": r["vehicle_no"],
            "qty": float(r["qty"] or 0),
            "rate": float(r["rate"] or 0),
            "total_amount": float(r["total_amount"] or 0)
        })

    conn.close()

    return jsonify(data)

@app.route("/api/transport-entry-report", methods=["GET"])
def api_transport_entry_report():

    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            entry_date,
            sl_no,
            transporter_name,
            challan_no,
            vehicle_no,
            slip_no,
            qty,
            rate,
            fuel_amount,
            cash_taken,
            total_amount
        FROM transport_entries
        ORDER BY id DESC
    """)

    rows = cur.fetchall()

    data = []

    for row in rows:

        data.append({
            "entry_date": row[0],
            "sl_no": row[1],
            "transporter_name": row[2],
            "challan_no": row[3],
            "vehicle_no": row[4],
            "slip_no": row[5],
            "qty": row[6],
            "rate": row[7],
            "fuel_amount": row[8],
            "cash_taken": row[9],
            "total_amount": row[10]
        })

    conn.close()

    return jsonify(data)

@app.route("/api/save-transport-entry", methods=["POST"])
def api_save_transport_entry():

    data = request.get_json()

    conn = get_pg_conn()
    cur = conn.cursor()

    qty = float(data.get("qty") or 0)
    rate = float(data.get("rate") or 0)
    cash_taken = float(data.get("cash_taken") or 0)

    fuel_amount = qty * rate
    total_amount = fuel_amount + cash_taken

    cur.execute("""
        INSERT INTO transport_entries (
            entry_date,
            sl_no,
            transporter_id,
            transporter_name,
            challan_no,
            vehicle_no,
            slip_no,
            qty,
            rate,
            fuel_amount,
            cash_taken,
            total_amount
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data.get("entry_date"),
        data.get("sl_no"),
        data.get("transporter_id"),
        data.get("transporter_name"),
        data.get("challan_no"),
        data.get("vehicle_no"),
        data.get("slip_no"),
        qty,
        rate,
        fuel_amount,
        cash_taken,
        total_amount
    ))

    conn.commit()
    conn.close()

    return jsonify({
        "status": "success",
        "message": "Transport entry saved"
    })

@app.route("/api/export-transport-excel")
def api_export_transport_excel():

    return redirect(
        url_for("export_party_transport_excel")
    )

@app.route("/api/export-transport-pdf")
def api_export_transport_pdf():

    return redirect(
        url_for("export_party_transport_pdf")
    )

if __name__ == "__main__":
    app.run(debug=True)